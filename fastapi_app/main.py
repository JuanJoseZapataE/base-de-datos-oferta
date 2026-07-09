# ...existing code...

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.encoders import jsonable_encoder
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import pandas as pd
import io
import os
import math
from datetime import datetime, date, time
from dateutil.relativedelta import relativedelta
import re
import unicodedata
import xml.etree.ElementTree as ET
from dotenv import load_dotenv
# Cargar .env desde la carpeta del paquete (asegura carga aunque el cwd sea el padre)
load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
from sqlalchemy import create_engine, text, bindparam
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter




# URL de la base de datos: editar o usar la variable de entorno DATABASE_URL
# Ejemplo: mysql+pymysql://root:password@localhost/sena_oferta
DATABASE_URL = os.getenv("DATABASE_URL", "mysql+pymysql://root@127.0.0.1/Oferta")

engine = create_engine(DATABASE_URL)
app = FastAPI(title="Importador Excel -> MySQL (sena_oferta)")



# Endpoint para traer todos los programas filtrados (sin paginación)
@app.get('/programas/all')
def programas_all(
    year: Optional[str] = None,
    municipio: Optional[str] = None,
    centro: Optional[str] = None,
    nivel: Optional[str] = None,
    estrategia: Optional[str] = None,
    convenio: Optional[str] = None,
    vigencia: Optional[str] = None,
    numero_ficha: Optional[int] = None,
    search: Optional[str] = None,
    solo_certificados: Optional[str] = None,
):
    clauses = []
    params: dict = {}
    if year is not None:
        years = [y.strip() for y in str(year).split(',') if y.strip()]
        if years:
            if len(years) == 1:
                clauses.append('YEAR(p.fecha_corte) = :year_0')
            else:
                in_keys = []
                for i, val in enumerate(years):
                    key = f'year_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(p.fecha_corte) IN (' + ','.join(in_keys) + ')')
            if 'year_0' not in params and years:
                params['year_0'] = int(years[0])
    if municipio:
        municipios = [m.strip().lower() for m in str(municipio).split(',') if m.strip()]
        if municipios:
            if len(municipios) == 1:
                clauses.append('LOWER(TRIM(p.ciudad_municipio)) = :municipio_0')
            else:
                in_keys = []
                for i, val in enumerate(municipios):
                    key = f'municipio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.ciudad_municipio)) IN (' + ','.join(in_keys) + ')')
            if 'municipio_0' not in params and municipios:
                params['municipio_0'] = municipios[0]
    if centro:
        centros = [c.strip().lower() for c in str(centro).split(',') if c.strip()]
        if centros:
            if len(centros) == 1:
                clauses.append('LOWER(TRIM(p.centro_formacion)) = :centro_0')
            else:
                in_keys = []
                for i, val in enumerate(centros):
                    key = f'centro_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.centro_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'centro_0' not in params and centros:
                params['centro_0'] = centros[0]
    if nivel:
        niveles = [n.strip().lower() for n in str(nivel).split(',') if n.strip()]
        if niveles:
            if len(niveles) == 1:
                clauses.append('LOWER(TRIM(p.nivel_formacion)) = :nivel_0')
            else:
                in_keys = []
                for i, val in enumerate(niveles):
                    key = f'nivel_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.nivel_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'nivel_0' not in params and niveles:
                params['nivel_0'] = niveles[0]
    if estrategia:
        estrategias = [e.strip().lower() for e in str(estrategia).split(',') if e.strip()]
        if estrategias:
            if len(estrategias) == 1:
                clauses.append('LOWER(TRIM(p.estrategia_programa)) = :estrategia_0')
            else:
                in_keys = []
                for i, val in enumerate(estrategias):
                    key = f'estrategia_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.estrategia_programa)) IN (' + ','.join(in_keys) + ')')
            if 'estrategia_0' not in params and estrategias:
                params['estrategia_0'] = estrategias[0]
    if convenio:
        convenios = [c.strip().lower() for c in str(convenio).split(',') if c.strip()]
        if convenios:
            if len(convenios) == 1:
                clauses.append('LOWER(TRIM(p.convenio)) = :convenio_0')
            else:
                in_keys = []
                for i, val in enumerate(convenios):
                    key = f'convenio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.convenio)) IN (' + ','.join(in_keys) + ')')
            if 'convenio_0' not in params and convenios:
                params['convenio_0'] = convenios[0]
    if vigencia is not None:
        vigencias = [v.strip() for v in str(vigencia).split(',') if v.strip()]
        if vigencias:
            if len(vigencias) == 1:
                clauses.append('YEAR(p.fecha_inicio) = :vigencia_0')
            else:
                in_keys = []
                for i, val in enumerate(vigencias):
                    key = f'vigencia_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(p.fecha_inicio) IN (' + ','.join(in_keys) + ')')
            if 'vigencia_0' not in params and vigencias:
                params['vigencia_0'] = int(vigencias[0])
    if numero_ficha is not None:
        clauses.append('p.numero_ficha = :numero_ficha')
        params['numero_ficha'] = int(numero_ficha)
    if search:
        s = str(search).strip().lower()
        if s:
            clauses.append('LOWER(TRIM(p.denominacion_programa)) LIKE :search')
            params['search'] = f'%{s}%'
    if solo_certificados and str(solo_certificados).strip().lower() not in {'0', 'false', 'no'}:
        clauses.append('(p.certificado IS NOT NULL AND p.certificado <> 0)')

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    sql = (
        get_programas_select_sql() +
        f'{where_sql} '
        'ORDER BY p.fecha_corte DESC, p.numero_ficha ASC, p.id ASC'
    )
    try:
        df = pd.read_sql(text(sql), con=engine, params=params)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error consultando programas: {e}')

    data = []
    if not df.empty:
        # Limpiar infinitos primero
        df = df.replace([float('inf'), float('-inf')], pd.NA)

        # Convertir columnas de fecha/tiempo a cadenas ISO para que sean JSON serializables
        for col in ['fecha_inicio',
                    'fecha_fin',
                    'fecha_corte',
                    'fecha_inicio_etapa_productiva',
                    'fecha_fin_etapa_productiva']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                df[col] = df[col].apply(
                    lambda v: v.isoformat() if hasattr(v, 'isoformat') else v
                )

        # Pasar a lista de dicts y reemplazar NaN/inf por None para que JSON los acepte
        raw_records = df.to_dict(orient='records')
        cleaned_records = []
        for row in raw_records:
            for key, value in list(row.items()):
                if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                    row[key] = None
            cleaned_records.append(row)
        data = cleaned_records

    return JSONResponse(data)




# Habilitar CORS para permitir peticiones desde el frontend local
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================
# FUNCIONES PARA REGISTRO CALIFICADO (inline para evitar problemas de importación)
# ============================================================================

def _read_excel_registro_calificado(content):
    """Lee un archivo Excel de registro calificado"""
    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, dtype=str)
        return df
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f'Error al leer el archivo Excel: {str(e)}'
        )


def _normalize_col_name(col_name: str) -> str:
    """Normaliza nombres de columna para búsqueda flexible (sin acentos)"""
    import unicodedata
    # Remover acentos
    text = str(col_name).lower().strip()
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    # Reemplazar espacios y guiones
    text = text.replace(' ', '_').replace('-', '_')
    return text


async def _process_registro_calificado(df: pd.DataFrame):
    """Procesa el DataFrame de registro calificado presencial e inserta en la tabla"""
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    # Normalizar nombres de columnas
    df.columns = [_normalize_col_name(col) for col in df.columns]

    # Mapear encabezados del Excel a nombres de columnas de la tabla
    column_mapping = {
        'proceso': 'proceso',
        'tipo_de_tramite': 'tipo_tramite',
        'tipo_tramite': 'tipo_tramite',  # Variante sin "de"
        'fecha_radicado': 'fecha_radicado',
        'numero_de_resolucion': 'numero_resolucion',
        'numero_resolucion': 'numero_resolucion',  # Variante sin "de"
        'fecha_de_resolucion': 'fecha_resolucion',
        'fecha_resolucion': 'fecha_resolucion',  # Variante sin "de"
        'resuelve': 'resuelve',
        'decreto_que_ampara': 'decreto_ampara',
        'decreto_ampara': 'decreto_ampara',
        'snies': 'snies',
        'cobertura': 'cobertura',
        'resolucion_ampara_el_programa': 'resolucion_ampara_programa',
        'resolucion_ampara_programa': 'resolucion_ampara_programa',
        'resolucion_ampara': 'resolucion_ampara',
        'resolucion_ampara_fecha': 'resolucion_ampara_fecha',
        'fecha_de_vencimiento': 'fecha_vencimiento',
        'fecha_vencimiento': 'fecha_vencimiento',  # Variante sin "de"
        'vigencia_rc': 'vigencia_rc',
        'cod_del_programa': 'cod_programa',
        'cod_programa': 'cod_programa',
        'version': 'version',
        'nombre_del_pro': 'nombre_programa',
        'programa': 'nombre_programa',
        'nombre_programa': 'nombre_programa',
        'nivel_de_formacion': 'nivel_formacion',
        'nivel_formacion': 'nivel_formacion',
        'red_de_conocimiento': 'red_conocimiento',
        'red_conocimiento': 'red_conocimiento',
        'modalidad': 'modalidad',
        'centro_de_formacion': 'centro_formacion',
        'centro_formacion': 'centro_formacion',
        'nombre_sede': 'nombre_sede',
        'tipo_sede': 'tipo_sede',
        'municipio': 'municipio',
        'lugar_de_desarrollo': 'lugar_desarrollo',
        'lugar_desarrollo': 'lugar_desarrollo',
        'direccion': 'direccion',
        'regional': 'regional',
        'nombre_regional': 'nombre_regional',
        'observaciones': 'observaciones',
        'clasificacion_para_tramite': 'clasificacion_tramite',
        'clasificacion_tramite': 'clasificacion_tramite',  # Variante sin "para"
        'aprendices_primer_cohorte': 'aprendices_primer_cohorte',
        'lugar_de_desarrollo_escrito_en_la_resolucion': 'lugar_desarrollo_resolucion',
    }

    # Crear DataFrame con nombres normalizados
    df_mapped = pd.DataFrame()
    for col in df.columns:
        if col in column_mapping:
            df_mapped[column_mapping[col]] = df[col]

    # Validar campos requeridos
    required_fields = ['proceso', 'tipo_tramite', 'numero_resolucion', 'nombre_sede', 'tipo_sede', 'clasificacion_tramite']
    missing_fields = [f for f in required_fields if f not in df_mapped.columns or df_mapped[f].isna().all()]
    
    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail=f'Faltan columnas requeridas: {", ".join(missing_fields)}'
        )

    # Limpiar datos: eliminar filas con campos requeridos vacíos
    for field in required_fields:
        if field in df_mapped.columns:
            df_mapped = df_mapped[df_mapped[field].notna()]
    
    df_mapped = df_mapped[df_mapped['proceso'].astype(str).str.strip() != '']
    
    if df_mapped.empty:
        raise HTTPException(status_code=400, detail='No hay registros válidos después de validar campos requeridos')

    try:
        with engine.connect() as conn:
            inserted_count = 0
            for idx, row in df_mapped.iterrows():
                # Preparar valores para inserción
                values = {}
                for col in df_mapped.columns:
                    val = row[col]
                    
                    # Convertir a None si es NaN
                    if pd.isna(val):
                        values[col] = None
                    # Convertir fechas
                    elif col in ['fecha_radicado', 'fecha_resolucion', 'resolucion_ampara_fecha', 'fecha_vencimiento']:
                        try:
                            parsed_date = pd.to_datetime(val)
                            values[col] = str(parsed_date.date()) if parsed_date else None
                        except:
                            values[col] = None
                    # Convertir enteros
                    elif col in ['snies', 'cod_programa', 'version', 'regional', 'aprendices_primer_cohorte']:
                        try:
                            values[col] = int(float(str(val))) if val else None
                        except:
                            values[col] = None
                    else:
                        values[col] = str(val).strip() if val else None

                # Construir SQL INSERT simples con todos los campos
                cols = list(values.keys())
                col_names = ', '.join([f'`{c}`' for c in cols])
                placeholders = ', '.join([f':{c}' for c in cols])
                
                # Campos que se pueden actualizar (no son parte de la PK)
                updateable_cols = [c for c in cols if c not in ['proceso', 'tipo_tramite', 'numero_resolucion', 'nombre_sede', 'tipo_sede', 'clasificacion_tramite']]
                update_clause = ', '.join([f'`{c}`=:{c}' for c in updateable_cols]) if updateable_cols else '`id`=`id`'
                
                insert_sql = f"""
                INSERT INTO registro_calificado_presencial ({col_names})
                VALUES ({placeholders})
                ON DUPLICATE KEY UPDATE {update_clause}
                """
                
                try:
                    conn.execute(text(insert_sql), values)
                    inserted_count += 1
                except Exception as row_error:
                    print(f"Error insertando fila {idx}: {row_error}")
                    print(f"Valores: {values}")
            
            conn.commit()

        return {
            'status': 'success',
            'message': f'Se cargaron {inserted_count} registros correctamente en registro_calificado_presencial',
            'rows_processed': inserted_count
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al guardar en base de datos: {str(e)}'
        )


async def _get_registro_calificado_data():
    """Obtiene los datos de registro_calificado_presencial de la base de datos"""
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT 
                    id, proceso, tipo_tramite, fecha_radicado, numero_resolucion, 
                    fecha_resolucion, resuelve, snies, cod_programa, version, 
                    nombre_programa, nivel_formacion, modalidad, 
                    nombre_sede, tipo_sede, municipio, centro_formacion,
                    clasificacion_tramite, aprendices_primer_cohorte, 
                    vigencia_rc, regional, nombre_regional, observaciones,
                    lugar_desarrollo, direccion, fecha_registro
                FROM registro_calificado_presencial
                ORDER BY fecha_registro DESC
                LIMIT 100
            """))
            rows = result.fetchall()
            return [dict(row._mapping) for row in rows]
    except Exception as e:
        return []


# ============================================================================
# ENDPOINTS PARA REGISTRO CALIFICADO
# ============================================================================

@app.post('/registro-calificado/upload-excel')
async def upload_registro_calificado(file: UploadFile = File(...)):
    """Sube un archivo Excel de Registro Calificado."""
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls o .xlsx')

    content = await file.read()
    
    try:
        df = _read_excel_registro_calificado(content)
        result = await _process_registro_calificado(df)
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al procesar archivo: {str(e)}')


@app.get('/registro-calificado/data')
async def get_registro_calificado_list():
    """Obtiene los datos cargados de Registro Calificado."""
    try:
        with engine.connect() as conn:
            # Primero verificar si la tabla existe
            check_table_sql = """
            SELECT COUNT(*) as count
            FROM information_schema.TABLES 
            WHERE TABLE_SCHEMA = DATABASE() 
            AND TABLE_NAME = 'registro_calificado_presencial'
            """
            table_exists = conn.execute(text(check_table_sql)).fetchone()
            
            if not table_exists or table_exists[0] == 0:
                return JSONResponse({'items': [], 'message': 'Tabla no existe aún'})
            
            # Obtener TODOS los datos de la tabla
            result = conn.execute(text("""
                SELECT 
                    proceso, tipo_tramite, fecha_radicado, numero_resolucion, 
                    fecha_resolucion, resuelve, decreto_ampara, snies, cobertura,
                    resolucion_ampara_programa, resolucion_ampara, resolucion_ampara_fecha,
                    fecha_vencimiento, vigencia_rc, cod_programa, version, 
                    nombre_programa, nivel_formacion, red_conocimiento, modalidad, 
                    centro_formacion, nombre_sede, tipo_sede, municipio, 
                    lugar_desarrollo, direccion, regional, nombre_regional, 
                    observaciones, clasificacion_tramite, aprendices_primer_cohorte,
                    lugar_desarrollo_resolucion, fecha_registro
                FROM registro_calificado_presencial
                ORDER BY fecha_registro DESC
                LIMIT 1000
            """))
            rows = result.fetchall()
            data = []
            for idx, row in enumerate(rows, 1):
                row_dict = dict(row._mapping)
                # Convertir TODAS las fechas/datetimes a string para JSON (evitar errores de serialización)
                for key in list(row_dict.keys()):
                    val = row_dict[key]
                    if val is not None and hasattr(val, 'isoformat'):
                        # Convertir cualquier date/datetime a ISO format string
                        row_dict[key] = val.isoformat()
                row_dict['id'] = idx  # Generar id secuencial
                data.append(row_dict)
            return JSONResponse(content=jsonable_encoder({'items': data, 'total': len(data)}))
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse({'items': [], 'error': str(e), 'detail': error_detail}, status_code=500)


# ============================================================================
# ENDPOINTS Y FUNCIONES PARA SEGUIMIENTO DE METAS / OFERTA
# ============================================================================


def _read_excel_oferta(content: bytes) -> pd.DataFrame:
    try:
        # Intento estándar con detección automática
        df = read_excel_with_header_detection(content)

        # Normalizar nombres de columna y comprobar si se detectó la columna de 'oferta'
        normalized = [normalize_col_name(c) for c in df.columns]
        if any('oferta' in c for c in normalized) or 'no_de_oferta' in normalized or 'no_de_oferta' in '_'.join(normalized):
            return df

        # Si no encontramos la columna, intentar leer usando la segunda fila (header=1)
        try:
            df2 = pd.read_excel(io.BytesIO(content), header=1, engine='openpyxl')
            normalized2 = [normalize_col_name(c) for c in df2.columns]
            if any('oferta' in c for c in normalized2) or 'no_de_oferta' in normalized2:
                print(f"DEBUG: read_excel_with header=1 detected columns: {normalized2}")
                return df2
        except Exception:
            pass

        # Como último recurso devolver el df original (se manejará error más arriba)
        print(f"DEBUG: columnas detectadas (fallback): {normalized}")
        return df
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Error al leer Excel de oferta: {e}')


def _normalize_verificado_to_front(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().upper()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    if s in {'SI', 'VERIFICADO'}:
        return 'VERIFICADO'
    if s in {'NO', 'NO VERIFICADO'}:
        return 'NO VERIFICADO'
    return str(value)


def _map_verificado_to_db(conn, value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    norm = _normalize_verificado_to_front(value)
    try:
        col = conn.execute(
            text(
                """
                SELECT COLUMN_TYPE
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = 'OFERTA_seguimiento_metas'
                  AND COLUMN_NAME = 'verificado'
                """
            )
        ).fetchone()
        col_type = str(col[0]).upper() if col and col[0] is not None else ''
    except Exception:
        col_type = ''

    if 'VERIFICADO' in col_type:
        return norm
    if 'SI' in unicodedata.normalize('NFKD', col_type).encode('ascii', 'ignore').decode('ascii'):
        return 'SÍ' if norm == 'VERIFICADO' else 'NO'
    return norm


def _build_cod_ver(codigo_programa: Optional[str], version_programa: Optional[object]) -> Optional[str]:
    codigo = clean_optional_text(codigo_programa)
    if codigo is None:
        return None

    version_value = version_programa
    try:
        if pd.isna(version_value):
            return None
    except Exception:
        pass

    if version_value is None:
        return None

    if isinstance(version_value, float):
        if version_value.is_integer():
            version_text = str(int(version_value))
        else:
            version_text = str(version_value).strip()
    else:
        version_text = str(version_value).strip()

    if version_text in {'', 'nan', 'nat', 'none', '<na>'}:
        return None

    return f'{codigo}-{version_text}'


def _check_registro_calificado_vigencia(conn, codigo_programa_str: Optional[str]) -> Optional[str]:
    """Verifica si codigo_programa existe en registro_calificado y si su fecha de vencimiento es vigente."""
    if codigo_programa_str is None or codigo_programa_str.strip() == '':
        return None

    try:
        codigo_prog_int = int(str(codigo_programa_str).strip())
    except (ValueError, TypeError):
        return None

    try:
        result = conn.execute(
            text(
                """
                SELECT fecha_vencimiento
                FROM registro_calificado_presencial
                WHERE cod_programa = :cod_programa
                LIMIT 1
                """
            ),
            {'cod_programa': codigo_prog_int}
        ).fetchone()

        if result is None:
            return None

        fecha_vencimiento = result[0]
        if fecha_vencimiento is None:
            return None

        from datetime import date
        today = date.today()
        if fecha_vencimiento > today:
            return 'VERIFICADO'
        else:
            return 'REGISTRO VENCIDO'

    except Exception:
        return None


async def _process_oferta(df: pd.DataFrame):
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    try:
        df = df.copy()

        # Si los encabezados están en una fila superior (por ejemplo fila 2), detectarlos.
        # Buscar en las primeras 3 filas si contienen tokens esperados como 'no de oferta'
        try:
            header_row_index = None
            scan_rows = min(3, len(df.index))
            for i in range(scan_rows):
                row_vals = [str(x).strip() if pd.notna(x) else '' for x in df.iloc[i].tolist()]
                normalized_row = [normalize_col_name(v) for v in row_vals]
                joined = ' '.join(normalized_row)
                if 'no_de_oferta' in joined or ('no' in joined and 'oferta' in joined):
                    header_row_index = i
                    break
            if header_row_index is not None:
                # usar esa fila como encabezado
                df.columns = df.iloc[header_row_index].tolist()
                df = df.iloc[header_row_index+1:].reset_index(drop=True)
                print(f"DEBUG: Encabezado detectado en fila {header_row_index+1}: {list(df.columns)}")
        except Exception:
            pass

        # Normalizar nombres de columna para el mapeo
        df.columns = [normalize_col_name(c) for c in df.columns]

        # Mapear por encabezados reales detectados en la fila 2, usando los nombres
        # exactos que aparecen en el Excel y dejando fallback por alias normalizados.
        aliases = {
            'codigo_centro': ['codigo_centro', 'cod_centro', 'codigo_de_centro'],
            'centro_formacion': ['centro_de_formacion', 'centro_formacion', 'nombre_sede'],
            'tipo_oferta': ['tipo_de_oferta', 'tipo_oferta'],
            'denominacion_formacion': ['1_denominacion_de_la_formacion', 'denominacion_formacion', 'denominacion_programa', 'nombre_programa'],
            'modalidad': ['2_modalidad', 'modalidad'],
            'codigo_programa': ['3_codigo_programa', 'codigo_programa', 'cod_programa'],
            'version_programa': ['4_version_del_programa', 'version_programa', 'version'],
            'resolucion_snies': ['5_no_resolucion,_fecha_y_codigo_snies', 'resolucion_snies', 'snies', 'resolucion'],
            'justificacion_oferta': ['6_justificacion_de_la_oferta_educativa', 'justificacion_oferta', 'justificacion'],
            'grupos': ['7_grupos', 'grupos', 'grupo'],
            'cupos': ['8_cupos', 'cupos', 'cupo'],
            'duracion_meses': ['9_duracion_del_programa_meses', 'duracion_meses', 'duracion', 'duracion_en_meses'],
            'municipio': ['10_municipio', 'municipio', 'municipio_formacion'],
            'sede': ['11_sede', 'sede', 'nombre_sede'],
            'codigo_indicativa': ['codigo_indicativa', 'cod_indicativa'],
            'horario_formacion': ['horario_formacion', 'horario'],
            'estrategia': ['estrategia', 'estrategia_programa'],
            'fecha_inicio': ['fecha_inicio', 'inicio'],
            'fecha_fin': ['fecha_fin', 'fin'],
            'oferta': ['no_de_oferta', 'no_oferta', 'oferta'],
            'verificado': ['verificado', 'verificacion'],
        }

        mapped = pd.DataFrame(index=df.index)
        for target, alias_list in aliases.items():
            src = get_first_existing_column(df, [normalize_col_name(a) for a in alias_list])
            mapped[target] = df[src] if src else None
        mapped['verificado'] = None

        int_cols = ['version_programa', 'grupos', 'cupos', 'duracion_meses', 'oferta']
        for col in int_cols:
            mapped[col] = pd.to_numeric(mapped[col], errors='coerce').astype('Int64')

        for col in ['fecha_inicio', 'fecha_fin']:
            mapped[col] = mapped[col].apply(_parse_excel_fecha_value)

        text_cols = [
            'codigo_centro', 'centro_formacion', 'tipo_oferta', 'denominacion_formacion', 'modalidad',
            'codigo_programa', 'resolucion_snies', 'justificacion_oferta', 'municipio', 'sede',
            'codigo_indicativa', 'horario_formacion', 'estrategia'
        ]
        for col in text_cols:
            mapped[col] = mapped[col].apply(clean_optional_text)

        mapped = mapped.where(pd.notna(mapped), None)
        # Algunas hojas conservan filas vacías dentro del rango usado; descartarlas
        # evita insertar registros fantasmas sin datos.
        mapped = mapped.dropna(how='all').reset_index(drop=True)
        if mapped.empty:
            raise HTTPException(status_code=400, detail='El Excel no contiene filas válidas')
        rows = mapped.to_dict(orient='records')

        # Si por alguna razón 'oferta' no quedó mapeada, forzar la columna detectada por nombre.
        if 'oferta' in mapped.columns and mapped['oferta'].isnull().all():
            src = get_first_existing_column(df, ['no_de_oferta', 'no_oferta', 'oferta'])
            if src:
                mapped['oferta'] = df[src]
                rows = mapped.to_dict(orient='records')

        with engine.connect() as conn:
            create_sql = """
            CREATE TABLE IF NOT EXISTS OFERTA_seguimiento_metas (
                id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                codigo_centro VARCHAR(10) NULL,
                centro_formacion VARCHAR(150) NULL,
                tipo_oferta VARCHAR(50) NULL,
                denominacion_formacion VARCHAR(255) NULL,
                modalidad VARCHAR(50) NULL,
                codigo_programa VARCHAR(20) NULL,
                version_programa INT NULL,
                resolucion_snies TEXT NULL,
                justificacion_oferta TEXT NULL,
                grupos INT NULL,
                cupos INT NULL,
                duracion_meses INT NULL,
                municipio VARCHAR(100) NULL,
                sede VARCHAR(255) NULL,
                codigo_indicativa VARCHAR(255) NULL,
                horario_formacion VARCHAR(150) NULL,
                estrategia VARCHAR(150) NULL,
                fecha_inicio DATE NULL,
                fecha_fin DATE NULL,
                oferta TINYINT NULL,
                verificado VARCHAR(30) NULL,
                fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
            conn.execute(text(create_sql))
            try:
                conn.execute(text('ALTER TABLE OFERTA_seguimiento_metas ADD COLUMN oferta TINYINT NULL'))
            except Exception:
                pass
            try:
                conn.execute(text('ALTER TABLE OFERTA_seguimiento_metas MODIFY COLUMN codigo_indicativa VARCHAR(255) NULL'))
            except Exception:
                pass
            try:
                conn.execute(text('ALTER TABLE OFERTA_seguimiento_metas MODIFY COLUMN verificado VARCHAR(30) NULL'))
            except Exception:
                pass

            catalogo_cod_ver = set()
            try:
                catalogo_result = conn.execute(
                    text(
                        """
                        SELECT TRIM(UPPER(cod_ver)) AS cod_ver
                        FROM catalogo
                        WHERE cod_ver IS NOT NULL AND TRIM(cod_ver) <> ''
                        """
                    )
                )
                catalogo_cod_ver = {
                    str(row[0]).strip().upper()
                    for row in catalogo_result.fetchall()
                    if row[0] is not None and str(row[0]).strip() != ''
                }
            except Exception:
                catalogo_cod_ver = set()

            inserted = 0
            for r in rows:
                # Primero verificar contra registro calificado
                reg_calificado_verif = _check_registro_calificado_vigencia(conn, r.get('codigo_programa'))
                if reg_calificado_verif:
                    verif_db = reg_calificado_verif
                else:
                    # Si no está en registro calificado, verificar con catálogo
                    cod_ver = _build_cod_ver(r.get('codigo_programa'), r.get('version_programa'))
                    if cod_ver and cod_ver.strip().upper() in catalogo_cod_ver:
                        verif_db = 'VERIFICADO'
                    else:
                        verif_db = 'VERIFICACION MANUAL'

                def _clean_mysql_value(value):
                    if value is None:
                        return None
                    try:
                        if pd.isna(value):
                            return None
                    except Exception:
                        pass
                    if isinstance(value, float):
                        if value.is_integer():
                            return int(value)
                        return value
                    if isinstance(value, str):
                        text_value = value.strip()
                        if text_value.lower() in {'nan', 'nat', 'none', ''}:
                            return None
                        if text_value.endswith('.0') and text_value[:-2].isdigit():
                            return text_value[:-2]
                        return text_value
                    return value

                def _clean_text_field(field_name, value):
                    text_value = _clean_mysql_value(value)
                    if text_value is None:
                        return None
                    if not isinstance(text_value, str):
                        return text_value

                    max_lengths = {
                        'codigo_centro': 10,
                        'centro_formacion': 150,
                        'tipo_oferta': 50,
                        'denominacion_formacion': 255,
                        'modalidad': 50,
                        'codigo_programa': 20,
                        'municipio': 100,
                        'sede': 255,
                        'codigo_indicativa': 255,
                        'horario_formacion': 150,
                        'estrategia': 150,
                    }
                    max_len = max_lengths.get(field_name)
                    if max_len is not None and len(text_value) > max_len:
                        text_value = text_value[:max_len]
                    return text_value

                # Asegurar que 'oferta' no sea NULL para evitar errores de integridad
                tmp_oferta = _clean_mysql_value(r.get('oferta'))
                try:
                    oferta_val = int(tmp_oferta) if tmp_oferta is not None else 0
                except Exception:
                    try:
                        oferta_val = int(float(tmp_oferta)) if tmp_oferta is not None else 0
                    except Exception:
                        oferta_val = 0

                params = {
                    'codigo_centro': _clean_text_field('codigo_centro', r.get('codigo_centro')),
                    'centro_formacion': _clean_text_field('centro_formacion', r.get('centro_formacion')),
                    'tipo_oferta': _clean_text_field('tipo_oferta', r.get('tipo_oferta')),
                    'denominacion_formacion': _clean_text_field('denominacion_formacion', r.get('denominacion_formacion')),
                    'modalidad': _clean_text_field('modalidad', r.get('modalidad')),
                    'codigo_programa': _clean_text_field('codigo_programa', r.get('codigo_programa')),
                    'version_programa': _clean_mysql_value(r.get('version_programa')),
                    'resolucion_snies': _clean_mysql_value(r.get('resolucion_snies')),
                    'justificacion_oferta': _clean_mysql_value(r.get('justificacion_oferta')),
                    'grupos': _clean_mysql_value(r.get('grupos')),
                    'cupos': _clean_mysql_value(r.get('cupos')),
                    'duracion_meses': _clean_mysql_value(r.get('duracion_meses')),
                    'municipio': _clean_text_field('municipio', r.get('municipio')),
                    'sede': _clean_text_field('sede', r.get('sede')),
                    'codigo_indicativa': _clean_text_field('codigo_indicativa', r.get('codigo_indicativa')),
                    'horario_formacion': _clean_text_field('horario_formacion', r.get('horario_formacion')),
                    'estrategia': _clean_text_field('estrategia', r.get('estrategia')),
                    'fecha_inicio': _clean_mysql_value(r.get('fecha_inicio')),
                    'fecha_fin': _clean_mysql_value(r.get('fecha_fin')),
                    'oferta': oferta_val,
                    'verificado': _clean_mysql_value(verif_db),
                }

                conn.execute(
                    text(
                        """
                        INSERT INTO OFERTA_seguimiento_metas (
                            codigo_centro, centro_formacion, tipo_oferta, denominacion_formacion, modalidad,
                            codigo_programa, version_programa, resolucion_snies, justificacion_oferta, grupos,
                            cupos, duracion_meses, municipio, sede, codigo_indicativa, horario_formacion,
                            estrategia, fecha_inicio, fecha_fin, oferta, verificado
                        ) VALUES (
                            :codigo_centro, :centro_formacion, :tipo_oferta, :denominacion_formacion, :modalidad,
                            :codigo_programa, :version_programa, :resolucion_snies, :justificacion_oferta, :grupos,
                            :cupos, :duracion_meses, :municipio, :sede, :codigo_indicativa, :horario_formacion,
                            :estrategia, :fecha_inicio, :fecha_fin, :oferta, :verificado
                        )
                        """
                    ),
                        params
                )
                inserted += 1

            conn.commit()

        return {
            'status': 'success',
            'message': f'Se insertaron {inserted} filas en OFERTA_seguimiento_metas',
            'rows_processed': inserted,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al guardar en OFERTA_seguimiento_metas: {e}')


async def _get_oferta_data(verificado_filter: Optional[str] = None):
    try:
        with engine.connect() as conn:
            check_sql = """
            SELECT COUNT(*) as count
            FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'OFERTA_seguimiento_metas'
            """
            exists = conn.execute(text(check_sql)).fetchone()
            if not exists or exists[0] == 0:
                return []

            # Construir WHERE clause si hay filtro
            where_clause = ''
            if verificado_filter:
                v = str(verificado_filter).strip().upper()
                if v in {'VERIFICADO', 'NO VERIFICADO', 'VERIFICACION MANUAL', 'REGISTRO VENCIDO'}:
                    where_clause = f"WHERE verificado = '{v}'"
            
            query = f"""
                SELECT id, codigo_centro, centro_formacion, tipo_oferta, denominacion_formacion,
                       modalidad, codigo_programa, version_programa, resolucion_snies,
                       justificacion_oferta, grupos, cupos, duracion_meses, municipio, sede,
                       codigo_indicativa, horario_formacion, estrategia, fecha_inicio,
                       fecha_fin, oferta, verificado, fecha_registro
                FROM OFERTA_seguimiento_metas
                {where_clause}
                ORDER BY fecha_registro DESC
                LIMIT 1000
            """
            
            result = conn.execute(text(query))
            out = []
            for r in result.fetchall():
                row = dict(r._mapping)
                # Convertir fechas a strings ISO
                if row.get('fecha_inicio'):
                    row['fecha_inicio'] = str(row['fecha_inicio'])
                if row.get('fecha_fin'):
                    row['fecha_fin'] = str(row['fecha_fin'])
                if row.get('fecha_registro'):
                    row['fecha_registro'] = str(row['fecha_registro'])
                row['verificado'] = _normalize_verificado_to_front(row.get('verificado'))
                out.append(row)
            return out
    except Exception as e:
        print(f"Error in _get_oferta_data: {e}")
        return []


async def _update_oferta_verificado(oferta_id: int, verificado_value: Optional[str]):
    try:
        with engine.connect() as conn:
            db_value = _map_verificado_to_db(conn, verificado_value)
            upd = conn.execute(
                text("UPDATE OFERTA_seguimiento_metas SET verificado = :v WHERE id = :id"),
                {'v': db_value, 'id': oferta_id},
            )
            conn.commit()
            return upd.rowcount
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error actualizando verificado: {e}')


@app.post('/seguimiento-metas/upload-oferta')
async def upload_oferta(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls o .xlsx')
    content = await file.read()
    df = _read_excel_oferta(content)
    result = await _process_oferta(df)
    return JSONResponse(result)


@app.get('/seguimiento-metas/data')
async def get_oferta_list(verificado: Optional[str] = None):
    data = await _get_oferta_data(verificado)
    return JSONResponse(content=jsonable_encoder({'items': data, 'total': len(data)}))


class VerificadoUpdateModel(BaseModel):
    id: int
    verificado: Optional[str] = None


@app.post('/seguimiento-metas/update-verificado')
async def post_update_verificado(payload: VerificadoUpdateModel):
    if payload.verificado is not None and payload.verificado not in {'VERIFICADO', 'NO VERIFICADO', 'VERIFICACION MANUAL', 'REGISTRO VENCIDO'}:
        raise HTTPException(status_code=400, detail='Valor de verificado inválido')
    count = await _update_oferta_verificado(payload.id, payload.verificado)
    return JSONResponse({'updated': count})


# ============================================================================
# ENDPOINTS PARA REGISTRO METAS INDIVIDUALES Y GRUPOS
# ============================================================================

def _ensure_registro_metas_tables():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS registro_metas_individuales (
                id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                tipo_formacion VARCHAR(50) NOT NULL,
                codigo_nivel_formacion VARCHAR(20) NOT NULL,
                codigo_programa_especial VARCHAR(20) NOT NULL,
                codigo_convenio VARCHAR(20) NOT NULL,
                tipo_modalidad VARCHAR(50) NOT NULL,
                nombre_meta VARCHAR(255) NOT NULL,
                meta_cupos INT NOT NULL,
                codigo_centro VARCHAR(20) NOT NULL,
                centro_formacion VARCHAR(255) NOT NULL,
                archivo_excel VARCHAR(500) NULL,
                grupo_id BIGINT NULL,
                fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_registro_metas_grupo (grupo_id),
                INDEX idx_registro_metas_centro (codigo_centro),
                INDEX idx_registro_metas_modalidad (tipo_modalidad)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS grupos_metas (
                id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                nombre_grupo VARCHAR(255) NOT NULL,
                total_cupos INT NOT NULL DEFAULT 0,
                cantidad_metas INT NOT NULL DEFAULT 0,
                archivo_consolidado VARCHAR(500) NULL,
                fecha_creacion DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_grupos_metas_fecha (fecha_creacion)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """))


class CrearGrupoMetasModel(BaseModel):
    nombre_grupo: str
    meta_ids: List[int]


@app.post('/registro-metas/crear')
async def crear_registro_meta(
    tipo_formacion: str = Form(...),
    codigo_nivel_formacion: str = Form(...),
    codigo_programa_especial: str = Form(...),
    codigo_convenio: str = Form(...),
    tipo_modalidad: str = Form(...),
    nombre_meta: str = Form(...),
    meta_cupos: str = Form(...),
    codigo_centro: str = Form(...),
    centro_formacion: str = Form(...),
):
    if tipo_formacion not in {'Titulada', 'Complementaria'}:
        raise HTTPException(status_code=400, detail='Tipo de formación inválido')
    if tipo_modalidad not in {'PRESENCIAL', 'A DISTANCIA', 'VIRTUAL'}:
        raise HTTPException(status_code=400, detail='Tipo modalidad inválido')
    if codigo_centro not in {'9308', '9121', '9223'}:
        raise HTTPException(status_code=400, detail='Centro de formación inválido')

    numeric_fields = {
        'codigo_nivel_formacion': codigo_nivel_formacion,
        'codigo_programa_especial': codigo_programa_especial,
        'codigo_convenio': codigo_convenio,
        'meta_cupos': meta_cupos,
    }
    for field_name, field_value in numeric_fields.items():
        if not str(field_value).strip().isdigit():
            raise HTTPException(status_code=400, detail=f'El campo {field_name} debe contener solo números')

    nombre_meta = nombre_meta.strip()
    if not nombre_meta:
        raise HTTPException(status_code=400, detail='El nombre de la meta es obligatorio')

    _ensure_registro_metas_tables()

    try:
        with engine.begin() as conn:
            result = conn.execute(text("""
                INSERT INTO registro_metas_individuales (
                    tipo_formacion, codigo_nivel_formacion, codigo_programa_especial,
                    codigo_convenio, tipo_modalidad, nombre_meta, meta_cupos,
                    codigo_centro, centro_formacion
                ) VALUES (
                    :tipo_formacion, :codigo_nivel_formacion, :codigo_programa_especial,
                    :codigo_convenio, :tipo_modalidad, :nombre_meta, :meta_cupos,
                    :codigo_centro, :centro_formacion
                )
            """), {
                'tipo_formacion': tipo_formacion,
                'codigo_nivel_formacion': codigo_nivel_formacion.strip(),
                'codigo_programa_especial': codigo_programa_especial.strip(),
                'codigo_convenio': codigo_convenio.strip(),
                'tipo_modalidad': tipo_modalidad,
                'nombre_meta': nombre_meta,
                'meta_cupos': int(meta_cupos),
                'codigo_centro': codigo_centro,
                'centro_formacion': centro_formacion.strip(),
            })
            new_id = result.lastrowid
        return JSONResponse({'id': new_id, 'message': 'Meta registrada correctamente'})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al registrar meta: {e}')


@app.get('/registro-metas/lista')
async def listar_registro_metas():
    _ensure_registro_metas_tables()
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT
                    id, tipo_formacion, codigo_nivel_formacion, codigo_programa_especial,
                    codigo_convenio, tipo_modalidad, nombre_meta, meta_cupos,
                    codigo_centro, centro_formacion, grupo_id, fecha_registro
                FROM registro_metas_individuales
                ORDER BY fecha_registro DESC, id DESC
            """)).fetchall()
            items = []
            for row in rows:
                item = dict(row._mapping)
                if item.get('fecha_registro') is not None and hasattr(item['fecha_registro'], 'isoformat'):
                    item['fecha_registro'] = item['fecha_registro'].isoformat(sep=' ', timespec='seconds')
                items.append(item)
            return JSONResponse(content=jsonable_encoder({'items': items, 'total': len(items)}))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al listar metas: {e}')


@app.post('/registro-metas/crear-grupo')
async def crear_grupo_metas(payload: CrearGrupoMetasModel):
    nombre_grupo = (payload.nombre_grupo or '').strip()
    if not nombre_grupo:
        raise HTTPException(status_code=400, detail='El nombre del grupo es obligatorio')
    if not payload.meta_ids:
        raise HTTPException(status_code=400, detail='Debe seleccionar al menos una meta')

    meta_ids = sorted({int(mid) for mid in payload.meta_ids if int(mid) > 0})
    if not meta_ids:
        raise HTTPException(status_code=400, detail='IDs de metas inválidos')

    _ensure_registro_metas_tables()

    try:
        with engine.begin() as conn:
            placeholders = ', '.join([f':id_{i}' for i in range(len(meta_ids))])
            params = {f'id_{i}': meta_id for i, meta_id in enumerate(meta_ids)}
            rows = conn.execute(text(f"""
                SELECT id, meta_cupos, grupo_id
                FROM registro_metas_individuales
                WHERE id IN ({placeholders})
            """), params).fetchall()

            if len(rows) != len(meta_ids):
                raise HTTPException(status_code=404, detail='Una o más metas seleccionadas no existen')

            for row in rows:
                if row.grupo_id is not None:
                    raise HTTPException(
                        status_code=400,
                        detail=f'La meta ID {row.id} ya pertenece al grupo {row.grupo_id}'
                    )

            total_cupos = sum(int(row.meta_cupos or 0) for row in rows)
            cantidad_metas = len(rows)

            grupo_result = conn.execute(text("""
                INSERT INTO grupos_metas (nombre_grupo, total_cupos, cantidad_metas)
                VALUES (:nombre_grupo, :total_cupos, :cantidad_metas)
            """), {
                'nombre_grupo': nombre_grupo,
                'total_cupos': total_cupos,
                'cantidad_metas': cantidad_metas,
            })
            grupo_id = grupo_result.lastrowid

            conn.execute(text(f"""
                UPDATE registro_metas_individuales
                SET grupo_id = :grupo_id
                WHERE id IN ({placeholders})
            """), {**params, 'grupo_id': grupo_id})

        return JSONResponse({
            'id': grupo_id,
            'nombre_grupo': nombre_grupo,
            'total_cupos': total_cupos,
            'cantidad_metas': cantidad_metas,
            'message': 'Grupo creado correctamente',
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al crear grupo: {e}')


@app.get('/registro-metas/grupos')
async def listar_grupos_metas():
    _ensure_registro_metas_tables()
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT id, nombre_grupo, total_cupos, cantidad_metas, fecha_creacion
                FROM grupos_metas
                ORDER BY fecha_creacion DESC, id DESC
            """)).fetchall()
            items = []
            for row in rows:
                item = dict(row._mapping)
                if item.get('fecha_creacion') is not None and hasattr(item['fecha_creacion'], 'isoformat'):
                    item['fecha_creacion'] = item['fecha_creacion'].isoformat(sep=' ', timespec='seconds')
                items.append(item)
            return JSONResponse(content=jsonable_encoder({'items': items, 'total': len(items)}))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al listar grupos: {e}')


@app.get('/')
def root():
    return {'message': 'API running. Usa /docs para ver los endpoints.'}

EXPECTED_COLUMNS = [
    'cod_regional', 'regional', 'cod_municipio', 'municipio', 'cod_centro', 'centro_formacion',
    'cod_programa', 'denominacion_programa', 'cod_ficha', 'estado_ficha', 'jornada', 'nivel_formacion',
    'cupo', 'inscritos_primera_opcion', 'inscritos_segunda_opcion', 'oferta', 'tipo', 'perfil_ingreso', 'periodo'
]

PROGRAMAS_COLUMNS = [
    'centro_formacion',
    'numero_ficha',
    'ciudad_municipio',
    'fecha_inicio',
    'fecha_fin',
    'nivel_formacion',
    'denominacion_programa',
    # Antes: estrato_programa. Ahora se usa como "estrategia del programa".
    'estrategia_programa',
    'convenio',
    'aprendices_activos',
    'certificado',
    'tipo_formacion',
    'modalidad_formacion',
    'estado_curso',
    'fecha_corte',
    'nombre_empresa',
    'aprendices_matriculados',
    'vigencia_aprendices',
    'fecha_inicio_etapa_productiva',
]

PROGRAMAS_SELECT_COLUMNS = [
    'id',
    'centro_formacion',
    'numero_ficha',
    'ciudad_municipio',
    'fecha_inicio',
    'fecha_fin',
    'nivel_formacion',
    'denominacion_programa',
    'estrategia_programa',
    'convenio',
    'aprendices_activos',
    'certificado',
    'tipo_formacion',
    'modalidad_formacion',
    'estado_curso',
    'fecha_corte',
    'nombre_empresa',
    'aprendices_matriculados',
    'vigencia_aprendices',
    'fecha_inicio_etapa_productiva',
]


def get_programas_select_sql() -> str:
    select_cols = ', '.join(f'p.{col}' for col in PROGRAMAS_SELECT_COLUMNS)
    return (
        f'SELECT {select_cols}, COALESCE(f.inscritos_primera_opcion, 0) AS inscritos '
        'FROM programas_formacion p '
        'LEFT JOIN fichas_formacion f ON f.cod_ficha = p.numero_ficha '
    )


INDICATIVA_COLUMNS = [
    'id_indicativa',
    'regional',
    'codigo_de_centro',
    'nombre_sede',
    'vigencia',
    'periodo_oferta',
    'codigo_programa',
    'version',
    'codigo_version',
    'nombre_programa',
    'nivel_de_formacion',
    'modalidad',
    'mes_inicio',
    'cupos',
    'ano_termina',
    'departamento_formacion',
    'codigo_dane_departamento',
    'municipio_formacion',
    'codigo_dane_municipio',
    'gira_tecnica',
    'programa_fic',
    'tipo_de_oferta',
    'persona_registra',
    'fecha_de_registro',
    'tipo_de_institucion',
    'nivel_institucion',
]

def ensure_programas_table():
    create_sql = """
    CREATE TABLE IF NOT EXISTS programas_formacion (
        id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
        centro_formacion VARCHAR(200) NULL,
        numero_ficha BIGINT NULL,
        ciudad_municipio VARCHAR(150) NULL,
        fecha_inicio DATE NULL,
        fecha_fin DATE NULL,
        nivel_formacion VARCHAR(100) NULL,
        denominacion_programa VARCHAR(255) NULL,
        estrategia_programa VARCHAR(255) NULL,
        convenio VARCHAR(255) NULL,
        aprendices_activos INT NULL,
        certificado VARCHAR(255) NULL,
        tipo_formacion VARCHAR(100) NULL,
        modalidad_formacion VARCHAR(100) NULL,
        estado_curso VARCHAR(100) NULL,
        fecha_corte DATE NULL,
        nombre_empresa VARCHAR(255) NULL,
        aprendices_matriculados INT NULL,
        vigencia_aprendices INT NULL,
        fecha_inicio_etapa_productiva DATE NULL,
        INDEX idx_programas_fecha_corte (fecha_corte),
        INDEX idx_programas_municipio (ciudad_municipio),
        INDEX idx_programas_numero_ficha (numero_ficha),
        INDEX idx_programas_vigencia (vigencia_aprendices)
    )
    """
    with engine.begin() as conn:
        conn.execute(text(create_sql))
        # Ajuste de esquema para instalaciones previas donde convenio quedo corto.
        try:
            conn.execute(text('ALTER TABLE programas_formacion MODIFY COLUMN convenio VARCHAR(255) NULL'))
        except Exception:
            pass
        # Migracion suave: instalaciones antiguas usaban estrato_programa.
        # Renombrar a estrategia_programa y ampliar longitud si existe.
        try:
            conn.execute(text('ALTER TABLE programas_formacion CHANGE COLUMN estrato_programa estrategia_programa VARCHAR(255) NULL'))
        except Exception:
            pass
        # Asegurar columna estado_curso para instalaciones previas.
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD COLUMN estado_curso VARCHAR(100) NULL'))
        except Exception:
            pass
        # Agregar nuevas columnas del PE_04 (modalidad_formacion, nombre_empresa, etc)
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD COLUMN modalidad_formacion VARCHAR(100) NULL'))
        except Exception:
            pass
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD COLUMN nombre_empresa VARCHAR(255) NULL'))
        except Exception:
            pass
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD COLUMN aprendices_matriculados INT NULL'))
        except Exception:
            pass
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD COLUMN vigencia_aprendices INT NULL'))
        except Exception:
            pass
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD COLUMN fecha_inicio_etapa_productiva DATE NULL'))
        except Exception:
            pass
        # Agregar índice en vigencia_aprendices si no existe
        try:
            conn.execute(text('ALTER TABLE programas_formacion ADD INDEX idx_programas_vigencia (vigencia_aprendices)'))
        except Exception:
            pass
        # inscritos ya no se guarda en ejecucion; se consulta desde fichas_formacion.inscritos_primera_opcion.
        try:
            conn.execute(text('ALTER TABLE programas_formacion DROP COLUMN cupos'))
        except Exception:
            pass


ensure_programas_table()


def ensure_indicativa_table():
    create_sql = """
    CREATE TABLE IF NOT EXISTS indicativa (
        id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
        id_indicativa BIGINT NULL,
        regional VARCHAR(150) NULL,
        codigo_de_centro INT NULL,
        nombre_sede VARCHAR(255) NULL,
        vigencia INT NULL,
        periodo_oferta VARCHAR(100) NULL,
        codigo_programa BIGINT NULL,
        version INT NULL,
        codigo_version VARCHAR(50) NULL,
        nombre_programa VARCHAR(255) NULL,
        nivel_de_formacion VARCHAR(150) NULL,
        modalidad VARCHAR(150) NULL,
        mes_inicio VARCHAR(50) NULL,
        cupos INT NULL,
        ano_termina INT NULL,
        departamento_formacion VARCHAR(150) NULL,
        codigo_dane_departamento VARCHAR(20) NULL,
        municipio_formacion VARCHAR(150) NULL,
        codigo_dane_municipio VARCHAR(20) NULL,
        gira_tecnica VARCHAR(50) NULL,
        programa_fic VARCHAR(50) NULL,
        tipo_de_oferta VARCHAR(150) NULL,
        persona_registra VARCHAR(150) NULL,
        fecha_de_registro DATETIME NULL,
        tipo_de_institucion VARCHAR(150) NULL,
        nivel_institucion VARCHAR(150) NULL,
        INDEX idx_indicativa_vigencia (vigencia),
        INDEX idx_indicativa_periodo (periodo_oferta),
        INDEX idx_indicativa_centro (nombre_sede)
    )
    """
    with engine.begin() as conn:
        conn.execute(text(create_sql))


ensure_indicativa_table()


def normalize_cols(cols):
    return [
        normalize_col_name(c) if isinstance(c, str) else c
        for c in cols
    ]


def normalize_col_name(value: str) -> str:
    if not isinstance(value, str):
        return value
    s = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    s = s.strip().lower()
    s = s.replace(' ', '_').replace('.', '').replace('-', '_').replace('/', '_')
    s = s.replace('(', '').replace(')', '')
    return s


def looks_like_expected_headers(columns) -> bool:
    normalized = set(normalize_cols(columns))
    expected = set(EXPECTED_COLUMNS)
    matches = len(normalized.intersection(expected))
    return matches >= 4 and ('cod_ficha' in normalized or 'cod_regional' in normalized)


def detect_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 30) -> Optional[int]:
    expected = set(EXPECTED_COLUMNS)
    scan_limit = min(max_scan_rows, len(df_raw.index))
    best_row = None
    best_score = 0

    for idx in range(scan_limit):
        row_values = [value for value in df_raw.iloc[idx].tolist() if pd.notna(value)]
        normalized_row = normalize_cols([str(value) for value in row_values])
        normalized_set = set(normalized_row)
        score = len(normalized_set.intersection(expected))

        if 'cod_ficha' in normalized_set:
            score += 2

        if score > best_score:
            best_score = score
            best_row = idx

    if best_row is not None and best_score >= 4:
        return int(best_row)
    return None


def read_excel_with_header_detection(content: bytes) -> pd.DataFrame:
    try:
        # Usar openpyxl en data_only=True para obtener el valor calculado de celdas con fórmulas.
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active

        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            raise HTTPException(status_code=400, detail='El Excel no contiene filas')

        df_raw = pd.DataFrame(raw_rows)

        # Intentar encontrar la fila de encabezado en las primeras filas.
        header_row = detect_header_row(df_raw)
        if header_row is None:
            header_row = 1 if len(df_raw.index) > 1 else 0

        headers = df_raw.iloc[header_row].tolist()
        data = df_raw.iloc[header_row + 1 :].reset_index(drop=True)
        data.columns = headers

        if looks_like_expected_headers(data.columns):
            return data

        # Si no coincide, devolver de todos modos la tabla ya con encabezado detectado.
        return data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'No se pudo leer el Excel: {e}')


def normalize_tipo(value: str) -> str:
    if not isinstance(value, str):
        return value
    v = value.strip().lower()
    if 'presencial' in v and ('distancia' in v or 'a distancia' in v):
        return 'PRESENCIAL Y A DISTANCIA'
    if 'presencial' in v:
        return 'PRESENCIAL'
    if 'distancia' in v or 'a distancia' in v:
        return 'A DISTANCIA'
    if 'virtual' in v:
        return 'VIRTUAL'
    return value.upper()


def normalize_oferta(value) -> str:
    if value is None:
        return None
    s = str(value).strip().upper()
    # Prefer to return a single-character code compatible with CHAR(1) in the DB
    # Map digits
    if s.isdigit():
        if s[-1] in '1234':
            return s[-1]
    # Map common roman numerals I..IV
    roman_map = {'I': '1', 'II': '2', 'III': '3', 'IV': '4'}
    if s in roman_map:
        return roman_map[s]
    # Map by keywords: VIRTUAL -> 4, PRESENCIAL or DISTANCIA -> 1
    if 'VIRTUAL' in s:
        return '4'
    if 'PRESENCIAL' in s or 'DISTANCIA' in s or 'A DISTANCIA' in s:
        return '1'
    # If contains a digit anywhere, take last digit
    for ch in reversed(s):
        if ch in '1234':
            return ch
    # Fallback: take first character (trimmed) to avoid length errors
    return s[0]


def export_header_label(column_name: str) -> str:
    """Convierte nombres técnicos a encabezados legibles para Excel."""
    if not column_name:
        return ''
    s = str(column_name).strip().replace('_', ' ')
    s = s.replace(' cod ', ' codigo ')
    if s.startswith('cod '):
        s = 'codigo ' + s[4:]
    if s == 'cod':
        s = 'codigo'
    return ' '.join(word.capitalize() for word in s.split())


def export_header_label_indicativa(column_name: str) -> str:
    if not column_name:
        return ''
    mapping = {
        'nombre_sede': 'Centro de formacion',
        'nivel_de_formacion': 'Nivel de formacion',
        'nombre_programa': 'Denominacion del programa',
        'periodo_oferta': 'Periodo oferta',
        'tipo_de_oferta': 'Tipo oferta',
    }
    if column_name in mapping:
        return mapping[column_name]
    return export_header_label(column_name)


def get_first_existing_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    existing = set(df.columns)
    for c in candidates:
        if c in existing:
            return c
    return None


def get_column_by_keywords(df: pd.DataFrame, keyword_groups: List[List[str]]) -> Optional[str]:
    """Busca una columna cuyo nombre contenga todos los tokens de algun grupo."""
    cols = [str(c) for c in df.columns]
    for group in keyword_groups:
        for col in cols:
            if all(token in col for token in group):
                return col
    return None


def detect_header_row_by_aliases(df_raw: pd.DataFrame, alias_pool: set[str], max_scan_rows: int = 30) -> Optional[int]:
    best_idx = None
    best_score = 0
    scan_limit = min(max_scan_rows, len(df_raw.index))

    for idx in range(scan_limit):
        row_values = [value for value in df_raw.iloc[idx].tolist() if pd.notna(value)]
        normalized_row = set(normalize_col_name(str(value)) for value in row_values)
        score = len(normalized_row.intersection(alias_pool))
        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx is not None and best_score >= 3:
        return int(best_idx)
    return None


def read_excel_with_custom_header_detection(content: bytes, alias_pool: set[str]) -> pd.DataFrame:
    try:
        df_default = read_excel_basic(content)
    except HTTPException:
        raise
    except Exception:
        df_default = pd.DataFrame()

    if not df_default.empty:
        normalized_default = set(normalize_cols(df_default.columns))
        if len(normalized_default.intersection(alias_pool)) >= 3:
            return df_default

    try:
        df_raw = read_excel_no_header(content)
        header_row = detect_header_row_by_aliases(df_raw, alias_pool)
        if header_row is not None:
            return read_excel_with_header_row(content, header_row)
    except Exception:
        pass

    if not df_default.empty:
        return df_default

    raise HTTPException(status_code=400, detail='No se pudo leer el Excel. Verifica que sea un archivo valido (.xls o .xlsx).')


def clean_optional_text(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s == '':
        return None
    if s.lower() in {'nan', 'none', 'null', 'nat', '<na>'}:
        return None
    return s


def read_excel_basic(content: bytes) -> pd.DataFrame:
    """Lee un Excel simple desde bytes.

    - Primero intenta como .xlsx con openpyxl.
    - Si falla y el backend intenta usar otro engine que no está instalado
      (por ejemplo xlrd para .xls), se devuelve un error 400 entendible
      en lugar de romper con 500 Internal Server Error.
    """
    # 1) Intentar siempre como .xlsx (openpyxl)
    try:
        return pd.read_excel(io.BytesIO(content), engine='openpyxl')
    except Exception as e_openpyxl:
        # 2) Fallback genérico de pandas. Si el archivo es .xls y está instalado xlrd,
        #    pandas usará ese engine de forma automática.
        try:
            return pd.read_excel(io.BytesIO(content))
        except ImportError:
            # Caso típico: archivo .xls pero xlrd no está instalado.
            raise HTTPException(
                status_code=400,
                detail=(
                    'No se pudo leer el Excel porque falta soporte para archivos .xls. '
                    'Vuelve a ejecutar la instalacion de requisitos para habilitarlo '
                    'o convierte el archivo a .xlsx.'
                ),
            )
        except Exception:
            # Si tampoco se puede leer aquí, reportar error de formato de archivo.
            raise HTTPException(
                status_code=400,
                detail='No se pudo leer el Excel. Verifica que sea un archivo de Excel valido (.xls o .xlsx).',
            ) from e_openpyxl


def read_excel_no_header(content: bytes) -> pd.DataFrame:
    try:
        return pd.read_excel(io.BytesIO(content), header=None, engine='openpyxl')
    except Exception:
        return pd.read_excel(io.BytesIO(content), header=None)


def read_spreadsheetml_xml(content: bytes) -> pd.DataFrame:
    """Lee un archivo XML de Excel 2003 (SpreadsheetML) como tabla.

    Extrae la primera hoja (Worksheet/Table) y construye un DataFrame usando
    la primera fila como encabezados y el resto como filas de datos.
    """
    try:
        root = ET.fromstring(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'XML de Excel invalido: {e}')

    ws = root.find('.//{*}Worksheet')
    if ws is None:
        raise HTTPException(status_code=400, detail='No se encontro ningun Worksheet en el XML de Excel.')

    table = ws.find('.//{*}Table')
    if table is None:
        raise HTTPException(status_code=400, detail='No se encontro ninguna tabla (Table) en el XML de Excel.')

    rows_raw = []
    for row in table.findall('.//{*}Row'):
        # Soportar celdas con atributo ss:Index (saltos de columnas).
        cells: list[str] = []
        col_pos = 0
        for cell in row.findall('.//{*}Cell'):
            idx_attr = None
            for attr_name, attr_val in cell.attrib.items():
                if attr_name.endswith('Index'):
                    idx_attr = attr_val
                    break
            if idx_attr is not None:
                try:
                    col_pos = int(idx_attr) - 1
                except Exception:
                    pass

            data_el = cell.find('.//{*}Data')
            text = '' if data_el is None or data_el.text is None else str(data_el.text)

            if len(cells) <= col_pos:
                cells.extend([''] * (col_pos - len(cells)))
                cells.append(text)
            else:
                cells[col_pos] = text

            col_pos += 1

        # Ignorar filas completamente vacias
        if any(val.strip() for val in cells):
            rows_raw.append(cells)

    if not rows_raw:
        return pd.DataFrame()

    # Detectar la fila que realmente contiene los encabezados (no el titulo tipo "PE-04_").
    # Usamos palabras clave tipicas de tus archivos: IDENTIFICADOR_FICHA, NUMERO_FICHA,
    # NOMBRE_PROGRAMA_FORMACION, etc.
    header_aliases = [
        'identificador_ficha', 'numero_ficha', 'n_ficha', 'codigo_ficha', 'cod_ficha',
        'nombre_programa_formacion', 'denominacion_programa',
        'centro_formacion', 'ciudad_municipio', 'nivel_formacion',
    ]
    alias_pool = set(normalize_col_name(a) for a in header_aliases)

    best_idx = 0
    best_score = -1
    scan_limit = min(40, len(rows_raw))
    for idx in range(scan_limit):
        row = rows_raw[idx]
        # normalizar cada celda como si fuera nombre de columna
        norm_cells = set(normalize_col_name(c) for c in row if c is not None)
        score = len(norm_cells.intersection(alias_pool))
        if score > best_score:
            best_score = score
            best_idx = idx

    header = rows_raw[best_idx]
    data_rows = rows_raw[best_idx + 1 :]
    num_cols = len(header)

    normalized_rows = []
    for r in data_rows:
        if len(r) < num_cols:
            r = r + [''] * (num_cols - len(r))
        elif len(r) > num_cols:
            r = r[:num_cols]
        normalized_rows.append(r)

    df = pd.DataFrame(normalized_rows, columns=header)
    return df


def read_excel_with_header_row(content: bytes, header_row: int) -> pd.DataFrame:
    try:
        return pd.read_excel(io.BytesIO(content), header=header_row, engine='openpyxl')
    except Exception:
        return pd.read_excel(io.BytesIO(content), header=header_row)


def extract_fecha_corte_from_filename(filename: str):
    """Extrae fecha de corte desde nombre tipo PE-04_20260306_15+55.xlsx -> 2026-03-06."""
    if not filename:
        return None
    match = re.search(r'(\d{8})', filename)
    if not match:
        return None
    raw = match.group(1)
    try:
        return datetime.strptime(raw, '%Y%m%d').date()
    except Exception:
        return None


def extract_fecha_corte_from_excel_content(content: bytes):
    """Intenta obtener fecha de corte desde el contenido del archivo Excel.

    Caso de uso: archivos que no traen fecha en el nombre pero si en una
    celda fija, por ejemplo A3 (fila 3, columna 1).
    """
    # Leer el libro con openpyxl para acceder directamente a la celda A3
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None

    try:
        ws = wb.active
    except Exception:
        return None

    try:
        cell = ws['A3']
    except Exception:
        return None

    value = cell.value
    if value is None:
        return None

    # 1) Si ya viene como datetime/fecha de Excel, usarla directamente.
    if isinstance(value, datetime):
        return value.date()

    # 2) Intentar parseo generico con pandas (por si es texto tipo "2022-12-01").
    s = str(value).strip()
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors='coerce')
    except Exception:
        ts = None
    if ts is not None and not pd.isna(ts):
        return ts.date()

    # 3) Buscar patrones numericos dentro del texto.
    #    Primero AAAAMMDD (8 digitos), luego AAAAMM (6 digitos).
    m8 = re.search(r'(\d{8})', s)
    if m8:
        raw = m8.group(1)
        try:
            dt = datetime.strptime(raw, '%Y%m%d')
            return dt.date()
        except Exception:
            pass

    m6 = re.search(r'(\d{6})', s)
    if m6:
        raw = m6.group(1)
        try:
            year = int(raw[:4])
            month = int(raw[4:6])
            dt = datetime(year, month, 1)
            return dt.date()
        except Exception:
            pass
    return None


def extract_fecha_corte_from_excel_b1(content: bytes):
    """Obtiene la fecha de corte desde la celda B1 del Excel."""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None

    try:
        ws = wb.active
    except Exception:
        return None

    try:
        value = ws['B1'].value
    except Exception:
        return None

    return _parse_excel_fecha_value(value)


def _parse_excel_fecha_value(value) -> Optional[date]:
    """Intenta convertir un valor de celda de Excel a date, asumiendo formato dia/mes/año cuando es texto."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    # Intentar formato explicito dd/mm/yyyy primero (requerido por el usuario)
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # Fallback generico usando pandas (acepta mas variantes)
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors='coerce')
        if ts is not None and not pd.isna(ts):
            return ts.date()
    except Exception:
        pass
    # Patrones numericos: ddmmyyyy o yyyymmdd dentro del texto
    m8 = re.search(r'(\d{8})', s)
    if m8:
        raw = m8.group(1)
        for fmt in ('%d%m%Y', '%Y%m%d'):
            try:
                return datetime.strptime(raw, fmt).date()
            except Exception:
                continue
    return None


def _parse_excel_hora_value(value) -> Optional[time]:
    """Intenta convertir un valor de celda de Excel a time (HH:MM[:SS])."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    s = str(value).strip()
    if not s:
        return None
    # Intentar formatos comunes de hora
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).time().replace(microsecond=0)
        except Exception:
            continue
    # Fallback: si viene como numero de Excel (fraccion del dia), intentar convertir
    try:
        # Excel almacena horas como fraccion del dia; multiplicar por 24 para horas
        num = float(s)
        total_seconds = int(round(num * 24 * 3600))
        hh = (total_seconds // 3600) % 24
        mm = (total_seconds % 3600) // 60
        ss = total_seconds % 60
        return time(hour=hh, minute=mm, second=ss)
    except Exception:
        return None


CATALOGO_COLUMNS = [
    'cod_ver',
    'prf_codigo',
    'prf_version',
    'tipo_de_formacion',
    'prf_denominacion',
    'nivel_de_formacion',
    'prf_duracion_maxima',
    'prf_dur_etapa_lectiva',
    'prf_dur_etapa_prod',
    'prf_fch_registro',
    'fecha_activo_en_ejecucion',
    'prf_edad_min_requerida',
    'prf_grado_min_requerido',
    'prf_descripcion_requisito',
    'prf_resolucion',
    'prf_fecha_resolucion',
    'prf_apoyo_fic',
    'prf_creditos',
    'prf_alineada',
    'linea_tecnologica',
    'red_tecnologica',
    'red_de_conocimiento',
    'modalidad',
    'apuestas_prioritarias',
    'fic',
    'tipo_permiso',
    'multiple_inscripcion',
    'indice',
    'ocupacion',
    'fecha_corte',
]

CATALOGO_COLUMN_ALIASES = {
    'cod_ver': ['cod_ver', 'codigo_version', 'codigo_de_version', 'cod_version', 'version_codigo'],
    'prf_codigo': ['prf_codigo', 'codigo_programa', 'codigo_prf', 'codigo_del_programa', 'prfcode'],
    'prf_version': ['prf_version', 'version', 'numero_version'],
    'tipo_de_formacion': ['tipo_de_formacion', 'tipo_formacion'],
    'prf_denominacion': ['prf_denominacion', 'denominacion_programa', 'nombre_programa', 'nombre_programa_formacion'],
    'nivel_de_formacion': ['nivel_de_formacion', 'nivel_formacion'],
    'prf_duracion_maxima': ['prf_duracion_maxima', 'duracion_maxima', 'prf_duracion_minima', 'duracion_minima'],
    'prf_dur_etapa_lectiva': ['prf_dur_etapa_lectiva', 'duracion_etapa_lectiva'],
    'prf_dur_etapa_prod': ['prf_dur_etapa_prod', 'duracion_etapa_productiva'],
    'prf_fch_registro': ['prf_fch_registro', 'fecha_registro', 'fch_registro'],
    'fecha_activo_en_ejecucion': ['fecha_activo_en_ejecucion', 'fecha_activa_en_ejecucion'],
    'prf_edad_min_requerida': ['prf_edad_min_requerida', 'edad_minima_requerida'],
    'prf_grado_min_requerido': ['prf_grado_min_requerido', 'grado_minimo_requerido'],
    'prf_descripcion_requisito': ['prf_descripcion_requisito', 'descripcion_requisito'],
    'prf_resolucion': ['prf_resolucion', 'resolucion'],
    'prf_fecha_resolucion': ['prf_fecha_resolucion', 'fecha_resolucion'],
    'prf_apoyo_fic': ['prf_apoyo_fic', 'apoyo_fic'],
    'prf_creditos': ['prf_creditos', 'creditos'],
    'prf_alineada': ['prf_alineada', 'alineada'],
    'linea_tecnologica': ['linea_tecnologica'],
    'red_tecnologica': ['red_tecnologica'],
    'red_de_conocimiento': ['red_de_conocimiento'],
    'modalidad': ['modalidad'],
    'apuestas_prioritarias': ['apuestas_prioritarias'],
    'fic': ['fic'],
    'tipo_permiso': ['tipo_permiso'],
    'multiple_inscripcion': ['multiple_inscripcion'],
    'indice': ['indice'],
    'ocupacion': ['ocupacion'],
    'fecha_corte': ['fecha_corte'],
}


def _load_catalogo_dataframe(content: bytes, filename: str) -> pd.DataFrame:
    alias_pool = set()
    for aliases in CATALOGO_COLUMN_ALIASES.values():
        for alias in aliases:
            alias_pool.add(normalize_col_name(alias))

    if filename.lower().endswith('.xml'):
        df = read_spreadsheetml_xml(content)
    else:
        df = read_excel_with_custom_header_detection(content, alias_pool)

    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    df = df.copy()
    df.columns = normalize_cols(df.columns)

    rename_map = {}
    for target, aliases in CATALOGO_COLUMN_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_col_name(alias)
            if alias_norm in df.columns and alias_norm != target:
                rename_map[alias_norm] = target
                break
    if rename_map:
        df = df.rename(columns=rename_map)

    if 'cod_ver' not in df.columns:
        codigo_col = get_first_existing_column(df, ['codigo_version', 'cod_version', 'version_codigo'])
        if codigo_col:
            df['cod_ver'] = df[codigo_col]

    for col in CATALOGO_COLUMNS:
        if col not in df.columns:
            df[col] = None

    df = df[CATALOGO_COLUMNS].copy()

    int_cols = [
        'prf_codigo', 'prf_version', 'prf_duracion_maxima', 'prf_dur_etapa_lectiva', 'prf_dur_etapa_prod',
        'prf_edad_min_requerida', 'prf_creditos', 'indice'
    ]
    for col in int_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')

    date_cols = ['prf_fch_registro', 'fecha_activo_en_ejecucion', 'prf_fecha_resolucion', 'fecha_corte']
    for col in date_cols:
        df[col] = df[col].apply(_parse_excel_fecha_value)

    text_cols = [c for c in CATALOGO_COLUMNS if c not in int_cols and c not in date_cols]
    for col in text_cols:
        df[col] = df[col].apply(clean_optional_text)

    df['cod_ver'] = df['cod_ver'].apply(clean_optional_text)

    # Convertir NaN a None para compatibilidad con MySQL/pymysql
    df = df.where(pd.notna(df), None)

    mask = df['cod_ver'].isna() | (df['cod_ver'].astype(str).str.strip() == '')
    if mask.any():
        derived = (
            df.loc[mask, 'prf_codigo'].astype('Int64').astype(str).replace('<NA>', '')
            + '_'
            + df.loc[mask, 'prf_version'].astype('Int64').astype(str).replace('<NA>', '')
        )
        derived = derived.str.strip('_')
        df.loc[mask, 'cod_ver'] = derived.where(derived != '_', None)

    df = df[df['cod_ver'].notna() & (df['cod_ver'].astype(str).str.strip() != '')].copy()
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene una columna cod_ver valida')

    df = df.drop_duplicates(subset=['cod_ver'], keep='last').copy()
    return df


def _upsert_catalogo_rows(df: pd.DataFrame) -> int:
    rows = df.to_dict(orient='records')
    if not rows:
        return 0

    # Convertir NaN/inf a None para compatibilidad con pymysql
    import math
    for row in rows:
        for key in list(row.keys()):
            val = row[key]
            if isinstance(val, float):
                if math.isnan(val) or math.isinf(val):
                    row[key] = None

    insert_columns = CATALOGO_COLUMNS
    insert_sql = (
        f"INSERT INTO catalogo ({', '.join(insert_columns)}) VALUES ({', '.join(f':{col}' for col in insert_columns)}) "
        f"ON DUPLICATE KEY UPDATE {', '.join(f'{col} = VALUES({col})' for col in insert_columns if col != 'cod_ver')}"
    )

    with engine.begin() as conn:
        result = conn.execute(text(insert_sql), rows)

    return int(result.rowcount or len(rows))


def extract_fecha_reporte_from_filename_fichas(filename: str) -> Optional[date]:
    """Extrae fecha de reporte desde nombre tipo CCX_17032026.xlsx o CCX_17-03-2026.xlsx.

    Estructura esperada: (siglas_centro)_(fecha_reporte)
    """
    if not filename:
        return None
    name = os.path.splitext(os.path.basename(filename))[0]
    parts = name.split('_')
    candidate = None
    if len(parts) >= 2:
        candidate = parts[1]
    else:
        # Si no hay guion bajo, buscar bloque de 8 digitos en todo el nombre
        m = re.search(r'(\d{8})', name)
        if m:
            candidate = m.group(1)
    if not candidate:
        return None
    s = str(candidate).strip()
    # Normalizar separadores
    s_norm = s.replace('-', '/').replace('.', '/').replace(' ', '/')
    # Intentar dd/mm/yyyy
    try:
        return datetime.strptime(s_norm, '%d/%m/%Y').date()
    except Exception:
        pass
    # Intentar ddmmyyyy (sin separadores)
    if re.fullmatch(r'\d{8}', s):
        for fmt in ('%d%m%Y', '%Y%m%d'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
    return None


def extract_fecha_hora_reporte_fichas(content: bytes, filename: str):
    """Obtiene fecha (B4) y hora (B5) del Excel de fichas o, si falta la fecha, del nombre del archivo.

    - B4: fecha de reporte en formato dia/mes/año (preferido).
    - B5: hora de reporte (HH:MM u hora de Excel).
    - Si B4 no tiene valor interpretable, se intenta extraer la fecha desde el nombre.
    """
    fecha: Optional[date] = None
    hora: Optional[time] = None

    # Primero intentar leer directamente desde el contenido del Excel
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        try:
            fecha_val = ws['B4'].value
            hora_val = ws['B5'].value
        except Exception:
            fecha_val = None
            hora_val = None
        fecha = _parse_excel_fecha_value(fecha_val)
        hora = _parse_excel_hora_value(hora_val)
    except Exception:
        # Si no se puede abrir el libro, se intentara solo por nombre
        pass

    # Si la fecha sigue sin definirse, usar nombre del archivo
    if fecha is None:
        fecha = extract_fecha_reporte_from_filename_fichas(filename or '')

    return fecha, hora


@app.post('/upload-excel')
async def upload_excel(file: UploadFile = File(...), periodo: Optional[int] = Form(None), oferta: Optional[str] = Form(None), tipo: Optional[str] = Form(None)):
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xml')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls, .xlsx o .xml')

    content = await file.read()

    # Extraer fecha y hora de reporte desde el Excel (B4/B5) o, si falta la fecha,
    # desde el nombre del archivo. Esto permite saber de que corte es el archivo
    # que se esta subiendo en el modulo de fichas.
    fecha_reporte, hora_reporte = extract_fecha_hora_reporte_fichas(content, file.filename or '')

    # Si es XML, intentar leerlo como tabla antes de aplicar la logica de deteccion
    # de encabezados pensada para Excel.
    if file.filename.lower().endswith('.xml'):
        try:
            df = pd.read_xml(io.BytesIO(content))
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f'No se pudo leer el XML como tabla: {e}',
            )
    else:
        df = read_excel_with_header_detection(content)

    # Normalizar nombres de columnas
    df.columns = normalize_cols(df.columns)

    # Preparar dataframe para insertar: asegurarnos de que existan todas las columnas
    # Si faltan columnas opcionales (oferta, tipo, perfil_ingreso) no impediremos la subida;
    # las creamos y las rellenamos desde los placeholders del formulario más abajo.
    df_to_insert = df.copy()
    # Si el Excel tiene exactamente el número esperado de columnas, asignar nombres por posición
    if df.shape[1] == len(EXPECTED_COLUMNS):
        df_to_insert.columns = EXPECTED_COLUMNS
    # Añadir las columnas faltantes (las inicializamos con None)
    for col in EXPECTED_COLUMNS:
        if col not in df_to_insert.columns:
            df_to_insert[col] = None
    # Reordenar columnas para consistencia
    df_to_insert = df_to_insert[EXPECTED_COLUMNS].copy()

    # Intentar convertir columnas numéricas; esto eliminará filas que sean en realidad
    # encabezados leídos como datos (p.ej. 'COD_REGIONAL') porque se convertirán a NaN.
    int_cols = ['cod_regional', 'cod_municipio', 'cod_centro', 'cod_programa', 'cod_ficha',
                'cupo', 'inscritos_primera_opcion', 'inscritos_segunda_opcion', 'periodo']
    for col in int_cols:
        if col in df_to_insert.columns:
            df_to_insert[col] = pd.to_numeric(df_to_insert[col], errors='coerce')

    # Quitar filas que no tengan un cod_ficha válido (clave primaria necesaria)
    if 'cod_ficha' in df_to_insert.columns:
        before = len(df_to_insert)
        df_to_insert = df_to_insert[df_to_insert['cod_ficha'].notna()].copy()
        removed = before - len(df_to_insert)
        if removed:
            print(f'Removed {removed} rows that looked like headers or had invalid cod_ficha')

    # Ahora convertir periodo y otras columnas a enteros donde aplique
    try:
        if 'periodo' in df_to_insert.columns:
            df_to_insert['periodo'] = pd.to_numeric(df_to_insert['periodo'], errors='coerce').astype('Int64')
    except Exception:
        pass

    # Normalizar columnas si vienen en el Excel
    if 'oferta' in df_to_insert.columns:
        df_to_insert['oferta'] = df_to_insert['oferta'].apply(lambda v: normalize_oferta(v) if pd.notna(v) else v)
    if 'tipo' in df_to_insert.columns:
        df_to_insert['tipo'] = df_to_insert['tipo'].apply(lambda v: normalize_tipo(v) if pd.notna(v) else v)

    # Comportamiento requerido:
    # - Si la columna existe y al menos una fila tiene valor, permitimos subir.
    #   - Si el formulario provee un valor, rellenamos los nulos con el valor del formulario.
    #   - Si el formulario no provee valor, dejamos nulos donde existan.
    # - Si la columna existe pero todas las filas están vacías, la tratamos como "ausente":
    #   requerimos el valor en el formulario para rellenar toda la columna.
    # - Si la columna no existe, requerimos el valor en el formulario.

    # periodo
    if 'periodo' in df_to_insert.columns:
        # ya intentamos convertir a numérico más arriba
        has_any = df_to_insert['periodo'].notna().any()
        if has_any:
            if periodo is not None:
                try:
                    periodo_val = int(periodo)
                    df_to_insert['periodo'] = df_to_insert['periodo'].fillna(periodo_val)
                except Exception:
                    raise HTTPException(status_code=400, detail='Periodo inválido')
            # si periodo no se provee, permitimos nulos tal como vienen en el Excel
        else:
            # columna existe pero vacía en todas las filas -> necesitamos formulario
            if periodo is None:
                raise HTTPException(status_code=400, detail='Periodo requerido (ni en Excel ni en el formulario)')
            try:
                periodo_val = int(periodo)
            except Exception:
                raise HTTPException(status_code=400, detail='Periodo inválido')
            df_to_insert['periodo'] = periodo_val
    else:
        # columna ausente
        if periodo is None:
            raise HTTPException(status_code=400, detail='Periodo requerido (ni en Excel ni en el formulario)')
        try:
            periodo_val = int(periodo)
        except Exception:
            raise HTTPException(status_code=400, detail='Periodo inválido')
        df_to_insert['periodo'] = periodo_val

    # oferta
    # Permitir que el Excel no tenga la columna 'oferta'. Si existe, rellenar nulos con el formulario cuando se provea.
    if 'oferta' in df_to_insert.columns:
        has_any = df_to_insert['oferta'].notna().any()
        if has_any:
            if oferta:
                oferta_norm = normalize_oferta(oferta)
                df_to_insert['oferta'] = df_to_insert['oferta'].fillna(oferta_norm)
            # si no se provee oferta en el formulario, dejamos nulos donde existan
        else:
            # columna existe pero vacía en todas las filas -> si el formulario tiene valor, usarlo, si no dejar nulos
            if oferta:
                oferta_norm = normalize_oferta(oferta)
                df_to_insert['oferta'] = oferta_norm
            else:
                df_to_insert['oferta'] = df_to_insert['oferta']
    else:
        # columna ausente -> crearla y rellenar con el valor del formulario si está, o dejar None
        oferta_norm = normalize_oferta(oferta) if oferta else None
        df_to_insert['oferta'] = oferta_norm

    # tipo
    # Permitir que el Excel no tenga la columna 'tipo'. Si existe, rellenar nulos con el formulario cuando se provea.
    if 'tipo' in df_to_insert.columns:
        has_any = df_to_insert['tipo'].notna().any()
        if has_any:
            if tipo:
                tipo_norm = normalize_tipo(tipo)
                df_to_insert['tipo'] = df_to_insert['tipo'].fillna(tipo_norm)
            # si no se provee tipo en el formulario, dejamos nulos donde existan
        else:
            if tipo:
                tipo_norm = normalize_tipo(tipo)
                df_to_insert['tipo'] = tipo_norm
            else:
                df_to_insert['tipo'] = df_to_insert['tipo']
    else:
        tipo_norm = normalize_tipo(tipo) if tipo else None
        df_to_insert['tipo'] = tipo_norm

    # perfil_ingreso: si la columna no existe simplemente crearla (se puede dejar vacía)
    if 'perfil_ingreso' not in df_to_insert.columns:
        df_to_insert['perfil_ingreso'] = None

    # Insertar en la base de datos usando pandas.to_sql (append)
    try:
        df_to_insert.to_sql('fichas_formacion', con=engine, if_exists='append', index=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al insertar en la base de datos: {e}')

    # Incluir en la respuesta la fecha y hora de reporte detectadas para
    # que el frontend pueda mostrar de que archivo/corte se trata.
    fecha_str = fecha_reporte.strftime('%d/%m/%Y') if isinstance(fecha_reporte, date) else None
    hora_str = hora_reporte.strftime('%H:%M:%S') if isinstance(hora_reporte, time) else None

    return JSONResponse({'inserted': len(df_to_insert), 'fecha_reporte': fecha_str, 'hora_reporte': hora_str})


@app.post('/indicativa/upload-excel')
async def upload_indicativa_excel(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xml')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls, .xlsx o .xml')

    content = await file.read()

    # Permitir XML, igual que en otros modulos
    if file.filename.lower().endswith('.xml'):
        try:
            df = pd.read_xml(io.BytesIO(content))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f'No se pudo leer el XML de indicativa como tabla: {e}')
    else:
        df = read_excel_basic(content)

    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    # Normalizar nombres de columnas (quita acentos, pasa a minusculas, reemplaza espacios por _)
    df.columns = normalize_cols(df.columns)

    # Asegurar todas las columnas esperadas
    df_to_insert = pd.DataFrame()
    for col in INDICATIVA_COLUMNS:
        if col in df.columns:
            df_to_insert[col] = df[col]
        else:
            df_to_insert[col] = None

    # Tipos basicos
    for col in ['codigo_de_centro', 'vigencia', 'version', 'cupos', 'ano_termina', 'codigo_programa', 'id_indicativa']:
        if col in df_to_insert.columns:
            df_to_insert[col] = pd.to_numeric(df_to_insert[col], errors='coerce').astype('Int64')

    # fecha_de_registro puede venir como texto o fecha de Excel
    if 'fecha_de_registro' in df_to_insert.columns:
        try:
            df_to_insert['fecha_de_registro'] = pd.to_datetime(df_to_insert['fecha_de_registro'], errors='coerce')
        except Exception:
            pass

    # Eliminar filas completamente vacias en campos clave basicos (nombre_sede y nombre_programa)
    key_fields = ['nombre_sede', 'nombre_programa']
    df_to_insert = df_to_insert[~df_to_insert[key_fields].isna().all(axis=1)].copy()
    if df_to_insert.empty:
        raise HTTPException(status_code=400, detail='No se encontraron filas validas para insertar en indicativa')

    try:
        df_to_insert.to_sql('indicativa', con=engine, if_exists='append', index=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al insertar en la tabla indicativa: {e}')

    return JSONResponse({'inserted': int(len(df_to_insert))})


@app.get('/fichas')
def get_fichas(
    periodo: Optional[int] = None,
    oferta: Optional[str] = None,
    tipo: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
):
    """Devuelve paginado: 50 por página por defecto. Respuesta JSON con items y metadatos.
    """
    # validar parámetros de paginación
    try:
        page = int(page)
    except Exception:
        page = 1
    try:
        per_page = int(per_page)
    except Exception:
        per_page = 50
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 1000:
        per_page = 50

    clauses = []
    params = {}
    if periodo is not None:
        clauses.append('periodo = :periodo')
        params['periodo'] = int(periodo)
    if oferta is not None:
        clauses.append('oferta = :oferta')
        params['oferta'] = normalize_oferta(oferta)
    if tipo is not None:
        clauses.append('UPPER(tipo) = :tipo')
        params['tipo'] = normalize_tipo(tipo)

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    count_sql = 'SELECT COUNT(*) AS total FROM fichas_formacion' + where_sql

    try:
        with engine.connect() as conn:
            total = conn.execute(text(count_sql), params).scalar() or 0
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error contando registros: {e}')

    offset = (page - 1) * per_page
    # Order by periodo asc, then oferta asc, then cod_ficha
    data_sql = f'SELECT * FROM fichas_formacion{where_sql} ORDER BY periodo ASC, oferta ASC, cod_ficha ASC LIMIT :limit OFFSET :offset'
    params2 = dict(params)
    params2['limit'] = per_page
    params2['offset'] = offset

    try:
        df = pd.read_sql(data_sql, con=engine, params=params2)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al leer la base de datos: {e}')

    return JSONResponse({
        'items': df.to_dict(orient='records'),
        'total': int(total),
        'page': page,
        'per_page': per_page,
    })


@app.delete('/fichas/{cod_ficha}')
def delete_ficha(cod_ficha: int):
    """Eliminar una ficha por su `cod_ficha`. Devuelve 204 si fue eliminada, 404 si no existe."""
    try:
        with engine.begin() as conn:
            # Verificar existencia antes de eliminar
            exists = conn.execute(text('SELECT COUNT(*) FROM fichas_formacion WHERE cod_ficha = :id'), {'id': int(cod_ficha)}).scalar() or 0
            if int(exists) == 0:
                raise HTTPException(status_code=404, detail='Ficha no encontrada')
            # Ejecutar borrado (commit al salir del context manager)
            conn.execute(text('DELETE FROM fichas_formacion WHERE cod_ficha = :id'), {'id': int(cod_ficha)})
            return JSONResponse(status_code=204, content={})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error eliminando ficha: {e}')
    # Si rowcount no indicó eliminación pero no hubo error, devolver 204
    return JSONResponse(status_code=204, content={})


@app.get('/fichas/count')
def fichas_count():
    """Endpoint diagnóstico: devuelve el total de filas y hasta 5 filas de ejemplo."""
    try:
        with engine.connect() as conn:
            total = conn.execute(text('SELECT COUNT(*) FROM fichas_formacion')).scalar() or 0
        sample = pd.read_sql('SELECT * FROM fichas_formacion LIMIT 5', con=engine)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al consultar la base de datos: {e}')

    return JSONResponse({'total': int(total), 'sample': sample.to_dict(orient='records')})


@app.get('/fichas/all')
def fichas_all():
    """Devuelve todos los registros de la tabla `fichas_formacion` sin paginación."""
    try:
        # Orden por periodo asc (años más antiguos primero), luego por oferta asc y cod_ficha
        df = pd.read_sql('SELECT * FROM fichas_formacion ORDER BY periodo ASC, oferta ASC, cod_ficha ASC', con=engine)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al leer la base de datos: {e}')

    return JSONResponse(df.to_dict(orient='records'))


@app.get('/indicativa')
def get_indicativa(
    page: int = 1,
    per_page: int = 50,
    centro: Optional[str] = None,
    nivel: Optional[str] = None,
    periodo_oferta: Optional[str] = None,
    municipio: Optional[str] = None,
    search: Optional[str] = None,
):
    """Listado paginado de la tabla indicativa para el frontend, con filtros opcionales."""
    try:
        page = int(page)
    except Exception:
        page = 1
    try:
        per_page = int(per_page)
    except Exception:
        per_page = 50
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 200:
        per_page = 50

    # Construir filtros
    clauses = []
    params: dict = {}
    if centro:
        centros = [c.strip().lower() for c in str(centro).split(',') if c.strip()]
        if centros:
            if len(centros) == 1:
                clauses.append('LOWER(TRIM(nombre_sede)) = :centro_0')
            else:
                in_keys = []
                for i, val in enumerate(centros):
                    key = f'centro_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(nombre_sede)) IN (' + ','.join(in_keys) + ')')
            if 'centro_0' not in params and centros:
                params['centro_0'] = centros[0]
    if nivel:
        niveles = [n.strip().lower() for n in str(nivel).split(',') if n.strip()]
        if niveles:
            if len(niveles) == 1:
                clauses.append('LOWER(TRIM(nivel_de_formacion)) = :nivel_0')
            else:
                in_keys = []
                for i, val in enumerate(niveles):
                    key = f'nivel_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(nivel_de_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'nivel_0' not in params and niveles:
                params['nivel_0'] = niveles[0]
    if periodo_oferta:
        periodos = [p.strip().lower() for p in str(periodo_oferta).split(',') if p.strip()]
        if periodos:
            if len(periodos) == 1:
                clauses.append('LOWER(TRIM(periodo_oferta)) = :periodo_oferta_0')
            else:
                in_keys = []
                for i, val in enumerate(periodos):
                    key = f'periodo_oferta_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(periodo_oferta)) IN (' + ','.join(in_keys) + ')')
            if 'periodo_oferta_0' not in params and periodos:
                params['periodo_oferta_0'] = periodos[0]
    if municipio:
        municipios = [m.strip().lower() for m in str(municipio).split(',') if m.strip()]
        if municipios:
            if len(municipios) == 1:
                clauses.append('LOWER(TRIM(municipio_formacion)) = :municipio_0')
            else:
                in_keys = []
                for i, val in enumerate(municipios):
                    key = f'municipio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(municipio_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'municipio_0' not in params and municipios:
                params['municipio_0'] = municipios[0]
    if search:
        s = str(search).strip().lower()
        if s:
            clauses.append('LOWER(TRIM(nombre_programa)) LIKE :search')
            params['search'] = f'%{s}%'

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    # Contar total
    count_sql = f'SELECT COUNT(*) FROM indicativa{where_sql}'
    try:
        with engine.connect() as conn:
            total = conn.execute(text(count_sql), params).scalar() or 0
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error contando registros de indicativa: {e}')

    offset = (page - 1) * per_page
    sql = (
        'SELECT id, nombre_sede, municipio_formacion, nivel_de_formacion, nombre_programa, '
        'periodo_oferta, tipo_de_oferta '
        'FROM indicativa'
        f'{where_sql} '
        'ORDER BY vigencia DESC, periodo_oferta ASC, nombre_sede ASC '
        'LIMIT :limit OFFSET :offset'
    )
    params_data = dict(params)
    params_data['limit'] = per_page
    params_data['offset'] = offset
    try:
        df = pd.read_sql(text(sql), con=engine, params=params_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error consultando indicativa: {e}')

    if not df.empty:
        df = df.replace([float('inf'), float('-inf')], pd.NA)
        df = df.where(pd.notna(df), None)

    items = df.to_dict(orient='records') if not df.empty else []

    # Renombrar claves para que ya vayan con los nombres que usara el frontend
    mapped_items = []
    for row in items:
        mapped_items.append(
            {
                'id': row.get('id'),
                'centro_formacion': row.get('nombre_sede'),
                'municipio_formacion': row.get('municipio_formacion'),
                'nivel_formacion': row.get('nivel_de_formacion'),
                'denominacion_programa': row.get('nombre_programa'),
                'periodo_oferta': row.get('periodo_oferta'),
                'tipo_oferta': row.get('tipo_de_oferta'),
            }
        )

    # Asegurar que no queden NaN/inf en los datos antes de serializar a JSON
    cleaned_items = []
    for row in mapped_items:
        for key, value in list(row.items()):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                row[key] = None
        cleaned_items.append(row)

    return JSONResponse(
        content=jsonable_encoder(
            {
                'items': cleaned_items,
                'total': int(total),
                'page': page,
                'per_page': per_page,
            }
        )
    )


@app.get('/indicativa/export')
def export_indicativa_excel(
    centro: Optional[str] = None,
    nivel: Optional[str] = None,
    periodo_oferta: Optional[str] = None,
    municipio: Optional[str] = None,
    search: Optional[str] = None,
):
    """Exporta Excel de la tabla indicativa respetando los filtros activos, con fecha de registro en B1."""
    clauses = []
    params: dict = {}

    if centro:
        centros = [c.strip().lower() for c in str(centro).split(',') if c.strip()]
        if centros:
            if len(centros) == 1:
                clauses.append('LOWER(TRIM(nombre_sede)) = :centro_0')
            else:
                in_keys = []
                for i, val in enumerate(centros):
                    key = f'centro_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(nombre_sede)) IN (' + ','.join(in_keys) + ')')
            if 'centro_0' not in params and centros:
                params['centro_0'] = centros[0]
    if nivel:
        niveles = [n.strip().lower() for n in str(nivel).split(',') if n.strip()]
        if niveles:
            if len(niveles) == 1:
                clauses.append('LOWER(TRIM(nivel_de_formacion)) = :nivel_0')
            else:
                in_keys = []
                for i, val in enumerate(niveles):
                    key = f'nivel_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(nivel_de_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'nivel_0' not in params and niveles:
                params['nivel_0'] = niveles[0]
    if periodo_oferta:
        periodos = [p.strip().lower() for p in str(periodo_oferta).split(',') if p.strip()]
        if periodos:
            if len(periodos) == 1:
                clauses.append('LOWER(TRIM(periodo_oferta)) = :periodo_oferta_0')
            else:
                in_keys = []
                for i, val in enumerate(periodos):
                    key = f'periodo_oferta_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(periodo_oferta)) IN (' + ','.join(in_keys) + ')')
            if 'periodo_oferta_0' not in params and periodos:
                params['periodo_oferta_0'] = periodos[0]
    if municipio:
        municipios = [m.strip().lower() for m in str(municipio).split(',') if m.strip()]
        if municipios:
            if len(municipios) == 1:
                clauses.append('LOWER(TRIM(municipio_formacion)) = :municipio_0')
            else:
                in_keys = []
                for i, val in enumerate(municipios):
                    key = f'municipio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(municipio_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'municipio_0' not in params and municipios:
                params['municipio_0'] = municipios[0]
    if search:
        s = str(search).strip().lower()
        if s:
            clauses.append('LOWER(TRIM(nombre_programa)) LIKE :search')
            params['search'] = f'%{s}%'

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    sql = (
        'SELECT * FROM indicativa'
        f'{where_sql} '
        'ORDER BY vigencia DESC, periodo_oferta ASC, nombre_sede ASC'
    )

    try:
        df = pd.read_sql(text(sql), con=engine, params=params)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al exportar indicativa: {e}')

    df_export = df.copy()
    
    # Detectar fecha de registro
    fecha_registro_value = None
    tiene_multiples_fechas = False
    
    if 'fecha_de_registro' in df_export.columns and not df_export.empty:
        fechas_unicas = df_export['fecha_de_registro'].dropna().dt.date.unique() if hasattr(df_export['fecha_de_registro'], 'dt') else pd.Series(df_export['fecha_de_registro'].dropna()).unique()
        if len(fechas_unicas) > 1:
            tiene_multiples_fechas = True
        elif len(fechas_unicas) == 1:
            fecha_registro_value = str(fechas_unicas[0])
    
    # Remover columna fecha_de_registro si no hay múltiples fechas
    if 'fecha_de_registro' in df_export.columns and not tiene_multiples_fechas:
        df_export = df_export.drop('fecha_de_registro', axis=1)

    # Encabezados legibles para Excel
    original_cols = list(df_export.columns)
    df_export.columns = [export_header_label_indicativa(col) for col in df_export.columns]

    # Columnas que SI se ven en el frontend
    visible_db_cols = {
        'nombre_sede',
        'nivel_de_formacion',
        'nombre_programa',
        'periodo_oferta',
        'tipo_de_oferta',
    }
    hidden_db_cols = {c for c in original_cols if c not in visible_db_cols}
    hidden_header_labels = {export_header_label_indicativa(c) for c in hidden_db_cols}

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir datos comenzando desde fila 3 si hay fecha_registro
        if fecha_registro_value:
            df_export.to_excel(writer, index=False, sheet_name='indicativa', startrow=2)
        else:
            df_export.to_excel(writer, index=False, sheet_name='indicativa')

        ws = writer.book['indicativa']
        
        # Agregar fecha de registro en B1
        if fecha_registro_value:
            ws['A1'] = 'Fecha de registro'
            ws['B1'] = fecha_registro_value
            ws['A1'].font = Font(bold=True, size=11)
            ws['B1'].font = Font(size=11)
        
        max_row = ws.max_row
        max_col = ws.max_column

        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = wrap_alignment

        # Verde success de Bootstrap para encabezados
        green_success = "198754"
        header_row = 3 if fecha_registro_value else 1
        for cell in ws[header_row]:
            if cell.value:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color=green_success, end_color=green_success, fill_type="solid")

        min_width = 12
        max_width = 60
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                cell_text = '' if value is None else str(value)
                if len(cell_text) > max_len:
                    max_len = len(cell_text)
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

            # Ocultar en Excel las columnas que no se ven en la tabla del frontend.
            header_value = ws.cell(row=header_row, column=col_idx).value
            if header_value in hidden_header_labels:
                ws.column_dimensions[col_letter].hidden = True
        
        # Ocultar columna fecha_de_registro si hay múltiples fechas
        if tiene_multiples_fechas and 'fecha_de_registro' in df.columns:
            fecha_col_idx = list(df.columns).index('fecha_de_registro')
            col_letter = get_column_letter(fecha_col_idx + 1)
            ws.column_dimensions[col_letter].hidden = True

    output.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'indicativa_export_{ts}.xlsx'

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.get('/indicativa/filters')
def get_indicativa_filters():
    """Devuelve valores distintos para los filtros de indicativa."""
    try:
        with engine.connect() as conn:
            centros = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT nombre_sede FROM indicativa WHERE nombre_sede IS NOT NULL ORDER BY nombre_sede ASC')
                ).fetchall()
                if r[0] is not None
            ]
            niveles = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT nivel_de_formacion FROM indicativa WHERE nivel_de_formacion IS NOT NULL ORDER BY nivel_de_formacion ASC')
                ).fetchall()
                if r[0] is not None
            ]
            periodos = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT periodo_oferta FROM indicativa WHERE periodo_oferta IS NOT NULL ORDER BY periodo_oferta ASC')
                ).fetchall()
                if r[0] is not None
            ]
            municipios = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT municipio_formacion FROM indicativa WHERE municipio_formacion IS NOT NULL ORDER BY municipio_formacion ASC')
                ).fetchall()
                if r[0] is not None
            ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error obteniendo filtros de indicativa: {e}')

    return JSONResponse(
        content=jsonable_encoder(
            {
                'centros': centros,
                'niveles': niveles,
                'periodos_oferta': periodos,
                'municipios': municipios,
            }
        )
    )


@app.delete('/indicativa/delete-all')
def delete_indicativa_all():
    """Elimina todos los registros de la tabla indicativa."""
    try:
        with engine.begin() as conn:
            result = conn.execute(text('DELETE FROM indicativa'))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error eliminando todos los registros de indicativa: {e}')

    return JSONResponse({'deleted_rows': int(result.rowcount or 0)})


@app.delete('/indicativa/{indicativa_id}')
def delete_indicativa_by_id(indicativa_id: int):
    """Elimina un registro de indicativa por su id."""
    try:
        with engine.begin() as conn:
            exists = conn.execute(
                text('SELECT COUNT(*) FROM indicativa WHERE id = :id'),
                {'id': int(indicativa_id)},
            ).scalar() or 0
            if int(exists) == 0:
                raise HTTPException(status_code=404, detail='Registro de indicativa no encontrado')

            result = conn.execute(
                text('DELETE FROM indicativa WHERE id = :id'),
                {'id': int(indicativa_id)},
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error eliminando registro de indicativa: {e}')

    return JSONResponse({'deleted_rows': int(result.rowcount or 0), 'id': int(indicativa_id)})


@app.get('/fichas/export')
def export_fichas_excel(
    centro: Optional[str] = None,
    oferta: Optional[str] = None,
    estado: Optional[str] = None,
    tipo: Optional[str] = None,
    nivel: Optional[str] = None,
    periodo: Optional[str] = None,
    search: Optional[str] = None,
):
    """Exporta Excel con los filtros activos (mismos criterios del frontend)."""
    clauses = []
    params = {}

    if centro:
        centros = [c.strip().lower() for c in str(centro).split(',') if c.strip()]
        if centros:
            if len(centros) == 1:
                clauses.append('LOWER(TRIM(centro_formacion)) = :centro_0')
            else:
                in_keys = []
                for i, val in enumerate(centros):
                    key = f'centro_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(centro_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'centro_0' not in params and centros:
                params['centro_0'] = centros[0]
    if oferta:
        ofertas = [normalize_oferta(o) for o in str(oferta).split(',') if str(o).strip()]
        if ofertas:
            if len(ofertas) == 1:
                clauses.append('oferta = :oferta_0')
            else:
                in_keys = []
                for i, val in enumerate(ofertas):
                    key = f'oferta_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('oferta IN (' + ','.join(in_keys) + ')')
            if 'oferta_0' not in params and ofertas:
                params['oferta_0'] = ofertas[0]
    if estado:
        estados = [e.strip().lower() for e in str(estado).split(',') if e.strip()]
        if estados:
            if len(estados) == 1:
                clauses.append('LOWER(TRIM(estado_ficha)) = :estado_0')
            else:
                in_keys = []
                for i, val in enumerate(estados):
                    key = f'estado_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(estado_ficha)) IN (' + ','.join(in_keys) + ')')
            if 'estado_0' not in params and estados:
                params['estado_0'] = estados[0]
    if tipo:
        tipos = [normalize_tipo(t).strip().lower() for t in str(tipo).split(',') if t.strip()]
        if tipos:
            if len(tipos) == 1:
                clauses.append('LOWER(TRIM(tipo)) = :tipo_0')
            else:
                in_keys = []
                for i, val in enumerate(tipos):
                    key = f'tipo_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(tipo)) IN (' + ','.join(in_keys) + ')')
            if 'tipo_0' not in params and tipos:
                params['tipo_0'] = tipos[0]
    if nivel:
        niveles = [n.strip().lower() for n in str(nivel).split(',') if n.strip()]
        if niveles:
            if len(niveles) == 1:
                clauses.append('LOWER(TRIM(nivel_formacion)) = :nivel_0')
            else:
                in_keys = []
                for i, val in enumerate(niveles):
                    key = f'nivel_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(nivel_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'nivel_0' not in params and niveles:
                params['nivel_0'] = niveles[0]
    if periodo is not None:
        periodos = [p.strip() for p in str(periodo).split(',') if p.strip()]
        if periodos:
            if len(periodos) == 1:
                clauses.append('periodo = :periodo_0')
            else:
                in_keys = []
                for i, val in enumerate(periodos):
                    key = f'periodo_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('periodo IN (' + ','.join(in_keys) + ')')
            if 'periodo_0' not in params and periodos:
                params['periodo_0'] = int(periodos[0])
    if search:
        clauses.append('LOWER(COALESCE(denominacion_programa, "")) LIKE :search')
        params['search'] = f"%{search.strip().lower()}%"

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    sql = f'SELECT * FROM fichas_formacion{where_sql} ORDER BY periodo ASC, oferta ASC, cod_ficha ASC'

    try:
        df = pd.read_sql(text(sql), con=engine, params=params)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al exportar desde la base de datos: {e}')

    # Exportar todas las columnas de la tabla fichas_formacion.
    df_export = df.copy()

    # Conservar los nombres originales para poder decidir que columnas ocultar.
    original_cols_fichas = list(df_export.columns)

    # Encabezados legibles para Excel: sin guion bajo y con formato titulo.
    df_export.columns = [export_header_label(col) for col in df_export.columns]

    # Columnas que NO se ven en la tabla del frontend (se ocultaran en Excel).
    hidden_fichas_db_cols = {'cod_municipio', 'cod_regional', 'cod_centro', 'perfil_ingreso'}
    hidden_fichas_headers = {
        export_header_label(c) for c in hidden_fichas_db_cols if c in original_cols_fichas
    }

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='fichas')

        ws = writer.book['fichas']
        max_row = ws.max_row
        max_col = ws.max_column

        # Ajuste de texto en todas las celdas (encabezado y datos).
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = wrap_alignment

        # Verde success de Bootstrap para encabezados
        green_success = "198754"
        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color=green_success, end_color=green_success, fill_type="solid")

        # Ajustar ancho por contenido: texto largo -> columna mas ancha, texto corto -> mas angosta.
        min_width = 12
        max_width = 60
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                cell_text = '' if value is None else str(value)
                if len(cell_text) > max_len:
                    max_len = len(cell_text)
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

            # Ocultar columnas que no se muestran en la tabla del frontend.
            header_value = ws.cell(row=1, column=col_idx).value
            if header_value in hidden_fichas_headers:
                ws.column_dimensions[col_letter].hidden = True

        # Crear una tabla de Excel para aplicar formato de tabla.
        if max_col >= 1 and max_row >= 1:
            last_col_letter = get_column_letter(max_col)
            table_ref = f'A1:{last_col_letter}{max_row}'
            table = Table(displayName='FichasExport', ref=table_ref)
            style = TableStyleInfo(
                name='TableStyleMedium9',
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)
    output.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'fichas_export_{ts}.xlsx'

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.post('/programas/upload-excel')
async def upload_programas_excel(
    file: UploadFile = File(...),
    fecha_corte_manual: Optional[date] = Form(None),
):
    """Subida normal de programas.

    Prioriza la fecha de corte manual enviada por el frontend.
    Si no se envia, usa compatibilidad con nombre del archivo o celda A3.
    """
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xml')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls, .xlsx o .xml')

    content = await file.read()
    fecha_corte_file = fecha_corte_manual
    if not fecha_corte_file:
        fecha_corte_file = extract_fecha_corte_from_filename(file.filename or '')
    # Si el nombre no trae fecha, intentar leerla desde el contenido (A3)
    if not fecha_corte_file and file.filename.lower().endswith(('.xls', '.xlsx')):
        fecha_corte_file = extract_fecha_corte_from_excel_content(content)
    if not fecha_corte_file:
        raise HTTPException(
            status_code=400,
            detail='No se pudo obtener fecha_corte. Envia fecha_corte_manual o usa un archivo con fecha en nombre/celda A3.',
        )

    stats = _process_programas_excel(content=content, filename=file.filename or '', fecha_corte_file=fecha_corte_file)
    return JSONResponse(stats)


@app.post('/programas/upload-excel-historico')
async def upload_programas_excel_historico(file: UploadFile = File(...), year: int = Form(...)):
    """Subida de archivos historicos de programas.

    Estos archivos no traen fecha de corte explicita, solo el anio. Se toma
    como fecha_corte el 31/12 de ese anio para que los filtros por anio
    funcionen igual que con los archivos normales.
    """
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xml')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls, .xlsx o .xml')

    try:
        year_int = int(year)
    except Exception:
        raise HTTPException(status_code=400, detail='El anio historico es invalido')

    if year_int < 1900 or year_int > 2100:
        raise HTTPException(status_code=400, detail='El anio historico debe estar entre 1900 y 2100')

    content = await file.read()
    # Usamos como fecha_corte el ultimo dia del anio para que YEAR(fecha_corte)
    # coincida con el anio historico que selecciona el usuario en los filtros.
    fecha_corte_file = date(year_int, 12, 31)

    stats = _process_programas_excel(content=content, filename=file.filename or '', fecha_corte_file=fecha_corte_file)
    return JSONResponse(stats)


def _process_programas_excel(*, content: bytes, filename: str, fecha_corte_file: date) -> dict:
    """Logica comun para importar programas desde un Excel.

    Se usa tanto para la subida normal como para la subida historica.
    """
    # Solo mapear los campos definidos para la tabla programas_formacion.
    # Se incluyen variantes que suelen venir en archivos tipo SOFIA/planeacion.
    col_map = {
        'centro_formacion': ['centro_formacion', 'centro_de_formacion', 'nombre_centro', 'nombre_centro_formacion'],
        'numero_ficha': ['numero_ficha', 'numero_de_ficha', 'n_ficha', 'codigo_ficha', 'cod_ficha', 'identificador_ficha'],
        'ciudad_municipio': ['ciudad_municipio', 'ciudad_o_municipio', 'nombre_ciudad', 'municipio', 'nombre_municipio_curso'],
        'fecha_inicio': ['fecha_inicio', 'fecha_de_inicio', 'inicio_ficha', 'fecha_inicio_ficha'],
        'fecha_fin': ['fecha_fin', 'fecha_de_fin', 'fin_ficha', 'fecha_fin_ficha', 'fecha_terminacion_ficha'],
        'nivel_formacion': ['nivel_formacion', 'nivel_de_formacion', 'nombre_nivel_formacion'],
        'denominacion_programa': ['denominacion_programa', 'denominacion_del_programa', 'nombre_curso', 'nombre_programa', 'nombre_programa_formacion'],
        # Nuevo uso: estrategia del programa; se alimenta principalmente desde
        # encabezados tipo NOMBRE_PROGRAMA_ESPECIAL.
        'estrategia_programa': [
            'estrategia_programa',
            'estrategia_del_programa',
            'nombre_programa_especial',
        ],
        # Estado del curso, viene tipicamente como ESTADO_CURSO.
        'estado_curso': [
            'estado_curso',
            'estado_del_curso',
        ],
        'convenio': ['convenio', 'nombre_convenio', 'tipo_convenio'],
        'aprendices_activos': ['aprendices_activos', 'total_aprendices_activos'],
        'aprendices_matriculados': [
            'aprendices_matriculados',
            'total_aprendices',
            'total_aprendices_matriculados',
            'matriculados',
            'aprendices_totales',
        ],
        'certificado': ['certificado'],
        # Ahora tipo_formacion se toma principalmente desde MODALIDAD_FORMACION
        # (normalizado a modalidad_formacion), manteniendo aliases anteriores
        # como compatibilidad por si vienen otros formatos viejos.
        'tipo_formacion': ['modalidad_formacion', 'tipo_formacion', 'tipo_de_formacion', 'nombre_tipo_formacion'],
        'modalidad_formacion': ['modalidad_formacion', 'modalidad_de_formacion', 'tipo_modalidad'],
        'nombre_empresa': ['nombre_empresa', 'empresa', 'razon_social'],
    }

    # Detectar si los encabezados reales no estan en la primera fila (caso tipico: primera fila con titulo PE-04_...)
    alias_pool = set()
    for aliases in col_map.values():
        for alias in aliases:
            alias_pool.add(normalize_col_name(alias))

    # Permitir XML: si la extension es .xml, leerlo como tabla antes de aplicar
    # la logica de deteccion de encabezados propia de Excel.
    if filename.lower().endswith('.xml'):
        # Para programas, los XML suelen ser archivos de Excel 2003 (SpreadsheetML).
        # Los leemos como hoja de calculo, no como tabla generica.
        try:
            df = read_spreadsheetml_xml(content)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f'No se pudo leer el XML de programas como Excel: {e}',
            )
    else:
        df = read_excel_basic(content)
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    norm_default_headers = [normalize_col_name(str(c)) for c in list(df.columns)]
    default_score = len(set(norm_default_headers).intersection(alias_pool))

    # Si detecta pocos encabezados utiles o muchos "unnamed", intenta encontrar la fila de encabezado correcta.
    unnamed_count = sum(1 for c in norm_default_headers if c.startswith('unnamed:'))
    if (default_score < 3 or unnamed_count >= 5) and not filename.lower().endswith('.xml'):
        # Solo intentamos la deteccion avanzada de fila de encabezado cuando el
        # archivo es realmente un Excel (xls/xlsx). Para XML ya tenemos el
        # DataFrame correcto desde pd.read_xml y reintentar leerlo como Excel
        # provoca errores de formato.
        df_raw = read_excel_no_header(content)
        scan_limit = min(40, len(df_raw.index))
        best_idx = None
        best_score = -1

        for idx in range(scan_limit):
            row_values = [str(v) for v in df_raw.iloc[idx].tolist() if pd.notna(v)]
            row_norm = set(normalize_cols(row_values))
            score = len(row_norm.intersection(alias_pool))
            if score > best_score:
                best_score = score
                best_idx = idx

        if best_idx is not None and best_score >= 3:
            df = read_excel_with_header_row(content, int(best_idx))

    df.columns = normalize_cols(df.columns)

    keyword_map = {
        'centro_formacion': [['centro'], ['nombre', 'centro']],
        'numero_ficha': [['ficha'], ['codigo', 'ficha']],
        'ciudad_municipio': [['ciudad'], ['municipio']],
        'fecha_inicio': [['fecha', 'inicio'], ['inicio']],
        'fecha_fin': [['fecha', 'fin'], ['fin']],
        'nivel_formacion': [['nivel', 'formacion'], ['nivel']],
        'denominacion_programa': [['nombre', 'curso'], ['denominacion', 'programa'], ['programa']],
        # Para estrategia_programa no usamos heuristica de palabras clave; se
        # confia en el mapeo explicito de col_map (NOMBRE_PROGRAMA_ESPECIAL).
        'convenio': [['convenio']],
        'aprendices_activos': [['aprendices', 'activos'], ['activos']],
        'certificado': [['certificado']],
        'tipo_formacion': [['tipo', 'formacion']],
    }

    df_out = pd.DataFrame()
    mapped_sources = {}
    for target in PROGRAMAS_COLUMNS:
        aliases = [normalize_col_name(a) for a in col_map.get(target, [target])]
        source_col = get_first_existing_column(df, aliases)
        if not source_col and target in keyword_map:
            source_col = get_column_by_keywords(df, keyword_map[target])
        if source_col:
            df_out[target] = df[source_col]
            mapped_sources[target] = source_col
        else:
            df_out[target] = None

    # fecha_corte se recibe ya calculada (normal o historica).
    df_out['fecha_corte'] = fecha_corte_file

    # Normalizacion de tipos
    for dcol in ['fecha_inicio', 'fecha_fin']:
        # En archivos historicos las fechas suelen venir como dia/mes/anio
        # (por ejemplo 31/12/2018), por eso usamos dayfirst=True.
        df_out[dcol] = pd.to_datetime(df_out[dcol], errors='coerce', dayfirst=True).dt.date

    for ncol in ['numero_ficha', 'aprendices_activos', 'aprendices_matriculados']:
        df_out[ncol] = pd.to_numeric(df_out[ncol], errors='coerce').astype('Int64')

    for scol in ['centro_formacion', 'ciudad_municipio', 'nivel_formacion', 'denominacion_programa', 'estrategia_programa', 'convenio', 'certificado', 'tipo_formacion', 'estado_curso', 'modalidad_formacion', 'nombre_empresa']:
        if scol in df_out.columns:
            df_out[scol] = df_out[scol].apply(clean_optional_text)
    
    # Crear columnas nuevas que se calcularan
    df_out['vigencia_aprendices'] = None
    df_out['fecha_inicio_etapa_productiva'] = None
    
    # Calcular vigencia_aprendices (año de fecha_inicio)
    if 'fecha_inicio' in df_out.columns:
        df_out['vigencia_aprendices'] = df_out['fecha_inicio'].apply(
            lambda x: x.year if pd.notna(x) and hasattr(x, 'year') else None
        )
    
    # Calcular fecha_inicio_etapa_productiva basada en nivel_formacion y fecha_fin
    
    def calc_fecha_etapa_productiva(row):
        """Calcula fecha de inicio de etapa productiva según nivel y fecha fin"""
        try:
            # row es una Serie de pandas cuando se usa axis=1
            fecha_fin = row['fecha_fin'] if 'fecha_fin' in row.index else None
            nivel = row['nivel_formacion'] if 'nivel_formacion' in row.index else None
            
            if pd.isna(fecha_fin) or fecha_fin is None:
                return None
            
            nivel_str = str(nivel).strip().upper() if nivel else ""
            
            if nivel_str in ['TECNÓLOGO', 'TECNICO', 'TÉCNICO']:
                return fecha_fin - relativedelta(months=6)
            elif nivel_str in ['OPERARIO', 'AUXILIAR']:
                return fecha_fin - relativedelta(months=3)
            else:
                return None
        except Exception as e:
            print(f"Error en calc_fecha_etapa_productiva: {e}")
            return None
    
    if 'fecha_fin' in df_out.columns and 'nivel_formacion' in df_out.columns:
        df_out['fecha_inicio_etapa_productiva'] = df_out.apply(calc_fecha_etapa_productiva, axis=1)

    total_rows_before_filter = len(df_out)
    num_ficha_with_value = int(df_out['numero_ficha'].notna().sum())
    denom_with_value = int(df_out['denominacion_programa'].fillna('').astype(str).str.strip().ne('').sum())

    # Eliminar filas completamente vacias en campos clave
    key_fields = ['numero_ficha', 'denominacion_programa']
    df_out = df_out[~df_out[key_fields].isna().all(axis=1)].copy()
    if df_out.empty:
        mapped_resume = ', '.join([f'{k}->{v}' for k, v in mapped_sources.items()]) if mapped_sources else 'ninguna'
        detected_headers = ', '.join([str(c) for c in list(df.columns)[:20]])
        raise HTTPException(
            status_code=400,
            detail=(
                'No se pudo insertar porque ninguna fila tiene datos en los campos clave '
                '(numero_ficha o denominacion_programa). '
                f'Filas leidas: {total_rows_before_filter}. '
                f'Filas con numero_ficha: {num_ficha_with_value}. '
                f'Filas con denominacion_programa: {denom_with_value}. '
                f'Mapeo detectado: {mapped_resume}. '
                f'Encabezados detectados (primeros 20): {detected_headers}'
            ),
        )

    # Validacion explicita para convenio (evita errores SQL opacos).
    convenio_max_len = 255
    if 'convenio' in df_out.columns:
        convenio_lengths = df_out['convenio'].dropna().astype(str).str.len()
        if not convenio_lengths.empty and int(convenio_lengths.max()) > convenio_max_len:
            bad_idx = convenio_lengths.idxmax()
            bad_value = str(df_out.loc[bad_idx, 'convenio'])
            preview = bad_value[:180]
            raise HTTPException(
                status_code=400,
                detail=(
                    f'El campo convenio supera el tamano permitido ({convenio_max_len}) en al menos una fila. '
                    f'Fila aproximada: {int(bad_idx) + 2}. '
                    f'Longitud encontrada: {int(convenio_lengths.max())}. '
                    f'Valor (preview): {preview}'
                ),
            )

    try:
        # Logica de persistencia:
        # - Si la ficha (numero_ficha) no existe aun en la tabla, se inserta
        #   el registro completo.
        # - Si la ficha ya existe, no se inserta una fila nueva; se actualizan
        #   algunos campos segun la fecha_corte:
        #   - aprendices_activos siempre se actualiza con el valor mas reciente
        #     disponible (si viene en el Excel).
        #   - "inscritos" ya no se guarda aqui; se consulta desde el modulo de
        #     inscripciones (fichas_formacion.inscritos_primera_opcion) por numero_ficha.
        #   - estrategia_programa y estado_curso se siguen actualizando de
        #     forma "suave" solo cuando vienen valores.

        # Normalizar lista de fichas del Excel (no nulas)
        ficha_series = df_out['numero_ficha'].dropna() if 'numero_ficha' in df_out.columns else pd.Series([], dtype='Int64')
        ficha_ids = [int(x) for x in ficha_series.tolist()]

        existing_ids: set[int] = set()
        existing_fechas: dict[int, Optional[date]] = {}
        if ficha_ids:
            check_sql = text('SELECT numero_ficha, fecha_corte FROM programas_formacion WHERE numero_ficha IN :ids').bindparams(bindparam('ids', expanding=True))
            with engine.connect() as conn:
                rows = conn.execute(check_sql, {'ids': ficha_ids}).fetchall()
            for r in rows:
                if not r or r[0] is None:
                    continue
                num = int(r[0])
                existing_ids.add(num)
                fecha_val = r[1] if len(r) > 1 else None
                if isinstance(fecha_val, datetime):
                    existing_fechas[num] = fecha_val.date()
                elif isinstance(fecha_val, date):
                    existing_fechas[num] = fecha_val
                else:
                    existing_fechas[num] = None

        # Estadisticas de duplicados (respecto a la tabla existente)
        duplicate_fichas = len(existing_ids)
        duplicate_rows_total = int(df_out['numero_ficha'].isin(existing_ids).sum()) if existing_ids and 'numero_ficha' in df_out.columns else 0

        # Filas nuevas (fichas que aun no existen en la tabla)
        df_new = df_out[~df_out['numero_ficha'].isin(existing_ids)].copy() if 'numero_ficha' in df_out.columns else df_out.copy()
        if not df_new.empty:
            df_new.to_sql('programas_formacion', con=engine, if_exists='append', index=False)

        # Filas existentes: actualizar segun reglas descritas arriba
        df_update = df_out[df_out['numero_ficha'].isin(existing_ids)].copy() if existing_ids else pd.DataFrame(columns=df_out.columns)

        updated_fichas = 0
        if not df_update.empty:
            update_params = []
            for _, row in df_update.iterrows():
                if pd.isna(row.get('numero_ficha')):
                    continue
                num_ficha = int(row['numero_ficha'])

                # aprendices_activos: siempre se actualiza si viene algun valor
                activos_val = row.get('aprendices_activos') if 'aprendices_activos' in df_update.columns else None
                activos_param = None
                if pd.notna(activos_val):
                    try:
                        activos_param = int(activos_val)
                    except Exception:
                        activos_param = None

                # estrategia_programa y estado_curso: comportamiento "suave"
                est_param = clean_optional_text(row['estrategia_programa']) if 'estrategia_programa' in df_update.columns else None
                estado_param = clean_optional_text(row['estado_curso']) if 'estado_curso' in df_update.columns else None

                update_params.append(
                    {
                        'numero_ficha': num_ficha,
                        'estrategia_programa': est_param,
                        'estado_curso': estado_param,
                        'aprendices_activos': activos_param,
                    }
                )

            if update_params:
                update_sql = text(
                    'UPDATE programas_formacion '
                    'SET '
                    '    estrategia_programa = COALESCE(:estrategia_programa, estrategia_programa), '
                    '    estado_curso = COALESCE(:estado_curso, estado_curso), '
                    '    aprendices_activos = COALESCE(:aprendices_activos, aprendices_activos) '
                    'WHERE numero_ficha = :numero_ficha'
                )
                with engine.begin() as conn:
                    result = conn.execute(update_sql, update_params)
                updated_fichas = int(result.rowcount or 0)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error insertando/actualizando programas: {e}')

    return {
        'inserted': int(len(df_new)) if 'df_new' in locals() else 0,
        'updated_fichas': int(updated_fichas),
        'duplicate_fichas': int(duplicate_fichas) if 'duplicate_fichas' in locals() else 0,
        'duplicate_rows': int(duplicate_rows_total) if 'duplicate_rows_total' in locals() else 0,
        'fecha_corte': str(fecha_corte_file),
    }


@app.post('/programas/upload-certificados')
async def upload_programas_certificados(file: UploadFile = File(...)):
    """Actualiza campo `certificado` en programas_formacion usando un Excel complementario y cruce por numero_ficha."""
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xml')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls, .xlsx o .xml')

    content = await file.read()
    is_xml = file.filename.lower().endswith('.xml')

    if is_xml:
        try:
            df = pd.read_xml(io.BytesIO(content))
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f'No se pudo leer el XML complementario de certificados como tabla: {e}',
            )
    else:
        df = read_excel_basic(content)
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel complementario no contiene filas')

    # Posibles encabezados para cruce y valor de certificados.
    ficha_aliases = [
        'numero_ficha', 'numero_de_ficha', 'identificador_ficha', 'codigo_ficha', 'cod_ficha', 'ficha'
    ]
    certificados_aliases = [
        'certificado', 'certificados', 'aprendices_certificados', 'total_aprendices_certificados', 'total_certificados'
    ]

    alias_pool = set(normalize_col_name(x) for x in ficha_aliases + certificados_aliases)
    norm_default_headers = [normalize_col_name(str(c)) for c in list(df.columns)]
    default_score = len(set(norm_default_headers).intersection(alias_pool))
    unnamed_count = sum(1 for c in norm_default_headers if c.startswith('unnamed:'))

    if (default_score < 2 or unnamed_count >= 5) and not is_xml:
        # Igual que en programas: solo intentamos la deteccion de encabezado
        # avanzada cuando el archivo es un Excel real. Para XML confiamos en
        # el DataFrame devuelto por pd.read_xml.
        df_raw = read_excel_no_header(content)
        scan_limit = min(40, len(df_raw.index))
        best_idx = None
        best_score = -1

        # Para certificados nos basta encontrar la fila donde aparezca
        # claramente la columna de ficha (por ejemplo, "Ficha"). Usamos
        # solo los aliases de ficha para que no dependa de tener tambien
        # encabezados de certificados en esa fila.
        ficha_alias_norm = set(normalize_col_name(x) for x in ficha_aliases)
        for idx in range(scan_limit):
            row_values = [str(v) for v in df_raw.iloc[idx].tolist() if pd.notna(v)]
            row_norm = set(normalize_cols(row_values))
            # puntuacion basada SOLO en coincidencias con encabezados de ficha
            score = len(row_norm.intersection(ficha_alias_norm))
            if score > best_score:
                best_score = score
                best_idx = idx
        # Si encontramos al menos una coincidencia con los aliases de ficha,
        # usamos esa fila como encabezado real.
        if best_idx is not None and best_score >= 1:
            df = read_excel_with_header_row(content, int(best_idx))

    df.columns = normalize_cols(df.columns)

    ficha_col = get_first_existing_column(df, [normalize_col_name(x) for x in ficha_aliases])
    cert_col = get_first_existing_column(df, [normalize_col_name(x) for x in certificados_aliases])

    # Se requiere SIEMPRE una columna de ficha. Si no existe, no podemos cruzar.
    if not ficha_col:
        headers = ', '.join([str(c) for c in list(df.columns)[:25]])
        raise HTTPException(
            status_code=400,
            detail=(
                'No se encontro ninguna columna de ficha en el Excel complementario. '
                f'Se esperaban encabezados similares a: {ficha_aliases}. '
                f'Encabezados detectados: {headers}'
            ),
        )

    # Si no hay columna explicita de cantidad de certificados, asumimos que
    # cada fila representa 1 certificado por ficha. Esto encaja con archivos
    # donde viene una fila por aprendiz certificado.
    if not cert_col:
        df_cert = pd.DataFrame({
            'numero_ficha': pd.to_numeric(df[ficha_col], errors='coerce').astype('Int64'),
            'certificado': 1,
        })
    else:
        df_cert = pd.DataFrame({
            'numero_ficha': pd.to_numeric(df[ficha_col], errors='coerce').astype('Int64'),
            'certificado': pd.to_numeric(df[cert_col], errors='coerce').astype('Int64'),
        })

    df_cert = df_cert[df_cert['numero_ficha'].notna()].copy()
    if df_cert.empty:
        raise HTTPException(status_code=400, detail='No hay filas con numero de ficha valido en el Excel complementario')

    # Si hay fichas repetidas en el archivo, sumar certificados para consolidar.
    df_cert = df_cert.groupby('numero_ficha', as_index=False)['certificado'].sum(min_count=1)
    df_cert['certificado'] = df_cert['certificado'].fillna(0).astype('Int64')

    ficha_ids = [int(x) for x in df_cert['numero_ficha'].dropna().tolist()]
    if not ficha_ids:
        raise HTTPException(status_code=400, detail='No se pudieron obtener fichas para actualizar')

    check_sql = text('SELECT DISTINCT numero_ficha FROM programas_formacion WHERE numero_ficha IN :ids').bindparams(bindparam('ids', expanding=True))
    with engine.connect() as conn:
        existing = conn.execute(check_sql, {'ids': ficha_ids}).fetchall()
    existing_ids = set(int(r[0]) for r in existing if r and r[0] is not None)

    df_to_update = df_cert[df_cert['numero_ficha'].isin(existing_ids)].copy()
    unmatched_ids = [fid for fid in ficha_ids if fid not in existing_ids]

    if df_to_update.empty:
        raise HTTPException(
            status_code=400,
            detail='Ninguna ficha del archivo complementario coincide con la tabla programas_formacion',
        )

    update_params = [
        {
            'numero_ficha': int(row['numero_ficha']),
            'certificado': str(int(row['certificado'])) if pd.notna(row['certificado']) else None,
        }
        for _, row in df_to_update.iterrows()
    ]

    update_sql = text('UPDATE programas_formacion SET certificado = :certificado WHERE numero_ficha = :numero_ficha')
    with engine.begin() as conn:
        result = conn.execute(update_sql, update_params)

    return JSONResponse({
        'updated_rows': int(result.rowcount or 0),
        'updated_fichas': int(len(df_to_update)),
        'unmatched_fichas': int(len(unmatched_ids)),
        'unmatched_sample': unmatched_ids[:20],
    })


@app.get('/programas')
def get_programas(
    year: Optional[str] = None,
    municipio: Optional[str] = None,
    centro: Optional[str] = None,
    nivel: Optional[str] = None,
    estrategia: Optional[str] = None,
    convenio: Optional[str] = None,
    vigencia: Optional[str] = None,
    numero_ficha: Optional[int] = None,
    search: Optional[str] = None,
    solo_certificados: Optional[str] = None,
    page: int = 1,
    per_page: int = 30,
):
    # Paginacion: maximo 30 registros por pagina
    try:
        page = int(page)
    except Exception:
        page = 1
    try:
        per_page = int(per_page)
    except Exception:
        per_page = 20
    if page < 1:
        page = 1
    if per_page < 1:
        per_page = 20
    if per_page > 20:
        per_page = 20

    clauses = []
    params: dict = {}

    if year is not None:
        years = [y.strip() for y in str(year).split(',') if y.strip()]
        if years:
            if len(years) == 1:
                clauses.append('YEAR(p.fecha_corte) = :year_0')
            else:
                in_keys = []
                for i, val in enumerate(years):
                    key = f'year_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(p.fecha_corte) IN (' + ','.join(in_keys) + ')')
            if 'year_0' not in params and years:
                params['year_0'] = int(years[0])
    if municipio:
        municipios = [m.strip().lower() for m in str(municipio).split(',') if m.strip()]
        if municipios:
            if len(municipios) == 1:
                clauses.append('LOWER(TRIM(p.ciudad_municipio)) = :municipio_0')
            else:
                in_keys = []
                for i, val in enumerate(municipios):
                    key = f'municipio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.ciudad_municipio)) IN (' + ','.join(in_keys) + ')')
            if 'municipio_0' not in params and municipios:
                params['municipio_0'] = municipios[0]
    if centro:
        centros = [c.strip().lower() for c in str(centro).split(',') if c.strip()]
        if centros:
            if len(centros) == 1:
                clauses.append('LOWER(TRIM(p.centro_formacion)) = :centro_0')
            else:
                in_keys = []
                for i, val in enumerate(centros):
                    key = f'centro_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.centro_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'centro_0' not in params and centros:
                params['centro_0'] = centros[0]
    if nivel:
        niveles = [n.strip().lower() for n in str(nivel).split(',') if n.strip()]
        if niveles:
            if len(niveles) == 1:
                clauses.append('LOWER(TRIM(p.nivel_formacion)) = :nivel_0')
            else:
                in_keys = []
                for i, val in enumerate(niveles):
                    key = f'nivel_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.nivel_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'nivel_0' not in params and niveles:
                params['nivel_0'] = niveles[0]
    if estrategia:
        estrategias = [e.strip().lower() for e in str(estrategia).split(',') if e.strip()]
        if estrategias:
            if len(estrategias) == 1:
                clauses.append('LOWER(TRIM(p.estrategia_programa)) = :estrategia_0')
            else:
                in_keys = []
                for i, val in enumerate(estrategias):
                    key = f'estrategia_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.estrategia_programa)) IN (' + ','.join(in_keys) + ')')
            if 'estrategia_0' not in params and estrategias:
                params['estrategia_0'] = estrategias[0]
    if convenio:
        convenios = [c.strip().lower() for c in str(convenio).split(',') if c.strip()]
        if convenios:
            if len(convenios) == 1:
                clauses.append('LOWER(TRIM(p.convenio)) = :convenio_0')
            else:
                in_keys = []
                for i, val in enumerate(convenios):
                    key = f'convenio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.convenio)) IN (' + ','.join(in_keys) + ')')
            if 'convenio_0' not in params and convenios:
                params['convenio_0'] = convenios[0]
    if vigencia is not None:
        vigencias = [v.strip() for v in str(vigencia).split(',') if v.strip()]
        if vigencias:
            if len(vigencias) == 1:
                clauses.append('YEAR(p.fecha_inicio) = :vigencia_0')
            else:
                in_keys = []
                for i, val in enumerate(vigencias):
                    key = f'vigencia_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(p.fecha_inicio) IN (' + ','.join(in_keys) + ')')
            if 'vigencia_0' not in params and vigencias:
                params['vigencia_0'] = int(vigencias[0])
    if numero_ficha is not None:
        clauses.append('p.numero_ficha = :numero_ficha')
        params['numero_ficha'] = int(numero_ficha)
    if search:
        s = str(search).strip().lower()
        if s:
            clauses.append('LOWER(TRIM(p.denominacion_programa)) LIKE :search')
            params['search'] = f'%{s}%'
    # solo_certificados: cualquier valor no vacio/"0"/"false" activa el filtro
    if solo_certificados and str(solo_certificados).strip().lower() not in {'0', 'false', 'no'}:
        clauses.append('(p.certificado IS NOT NULL AND p.certificado <> 0)')

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)


    count_sql = f'SELECT COUNT(*) AS total FROM programas_formacion p{where_sql}'
    try:
        with engine.connect() as conn:
            total = conn.execute(text(count_sql), params).scalar() or 0
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error contando programas: {e}')

    offset = (page - 1) * per_page
    data_sql = (
        get_programas_select_sql() +
        f'{where_sql} '
        'ORDER BY p.fecha_corte DESC, p.numero_ficha ASC, p.id ASC '
        'LIMIT :limit OFFSET :offset'
    )
    params_data = dict(params)
    params_data['limit'] = per_page
    params_data['offset'] = offset

    try:
        df = pd.read_sql(text(data_sql), con=engine, params=params_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error consultando programas: {e}')

    # JSON no soporta NaN/NaT/inf; convertir a None para serializar correctamente.
    if not df.empty:
        df = df.replace([float('inf'), float('-inf')], pd.NA)
        df = df.where(pd.notna(df), None)

    fecha_corte = None
    if not df.empty and 'fecha_corte' in df.columns:
        valid = pd.to_datetime(df['fecha_corte'], errors='coerce').dropna()
        if not valid.empty:
            fecha_corte = valid.max().date().isoformat()

    items = df.to_dict(orient='records')
    for row in items:
        for key, value in list(row.items()):
            try:
                if pd.isna(value):
                    row[key] = None
                    continue
            except Exception:
                pass
            if hasattr(value, 'isoformat') and not isinstance(value, str):
                try:
                    row[key] = value.isoformat()
                except Exception:
                    pass

    payload = {
        'items': items,
        'total': int(total),
        'fecha_corte': fecha_corte,
        'page': page,
        'per_page': per_page,
    }
    return JSONResponse(content=jsonable_encoder(payload))


@app.get('/programas/export')
def export_programas_excel(
    year: Optional[str] = None,
    municipio: Optional[str] = None,
    centro: Optional[str] = None,
    estrategia: Optional[str] = None,
    convenio: Optional[str] = None,
    vigencia: Optional[str] = None,
    numero_ficha: Optional[int] = None,
    search: Optional[str] = None,
    solo_certificados: Optional[str] = None,
):
    """Exporta Excel de programas_formacion respetando los filtros activos."""
    clauses = []
    params: dict = {}

    if year is not None:
        years = [y.strip() for y in str(year).split(',') if y.strip()]
        if years:
            if len(years) == 1:
                clauses.append('YEAR(p.fecha_corte) = :year_0')
            else:
                in_keys = []
                for i, val in enumerate(years):
                    key = f'year_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(p.fecha_corte) IN (' + ','.join(in_keys) + ')')
            if 'year_0' not in params and years:
                params['year_0'] = int(years[0])
    if municipio:
        municipios = [m.strip().lower() for m in str(municipio).split(',') if m.strip()]
        if municipios:
            if len(municipios) == 1:
                clauses.append('LOWER(TRIM(p.ciudad_municipio)) = :municipio_0')
            else:
                in_keys = []
                for i, val in enumerate(municipios):
                    key = f'municipio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.ciudad_municipio)) IN (' + ','.join(in_keys) + ')')
            if 'municipio_0' not in params and municipios:
                params['municipio_0'] = municipios[0]
    if centro:
        centros = [c.strip().lower() for c in str(centro).split(',') if c.strip()]
        if centros:
            if len(centros) == 1:
                clauses.append('LOWER(TRIM(p.centro_formacion)) = :centro_0')
            else:
                in_keys = []
                for i, val in enumerate(centros):
                    key = f'centro_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.centro_formacion)) IN (' + ','.join(in_keys) + ')')
            if 'centro_0' not in params and centros:
                params['centro_0'] = centros[0]
    if estrategia:
        estrategias = [e.strip().lower() for e in str(estrategia).split(',') if e.strip()]
        if estrategias:
            if len(estrategias) == 1:
                clauses.append('LOWER(TRIM(p.estrategia_programa)) = :estrategia_0')
            else:
                in_keys = []
                for i, val in enumerate(estrategias):
                    key = f'estrategia_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.estrategia_programa)) IN (' + ','.join(in_keys) + ')')
            if 'estrategia_0' not in params and estrategias:
                params['estrategia_0'] = estrategias[0]
    if convenio:
        convenios = [c.strip().lower() for c in str(convenio).split(',') if c.strip()]
        if convenios:
            if len(convenios) == 1:
                clauses.append('LOWER(TRIM(p.convenio)) = :convenio_0')
            else:
                in_keys = []
                for i, val in enumerate(convenios):
                    key = f'convenio_{i}'
                    in_keys.append(f':{key}')
                    params[key] = val
                clauses.append('LOWER(TRIM(p.convenio)) IN (' + ','.join(in_keys) + ')')
            if 'convenio_0' not in params and convenios:
                params['convenio_0'] = convenios[0]
    if vigencia is not None:
        vigencias = [v.strip() for v in str(vigencia).split(',') if v.strip()]
        if vigencias:
            if len(vigencias) == 1:
                clauses.append('YEAR(p.fecha_inicio) = :vigencia_0')
            else:
                in_keys = []
                for i, val in enumerate(vigencias):
                    key = f'vigencia_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(p.fecha_inicio) IN (' + ','.join(in_keys) + ')')
            if 'vigencia_0' not in params and vigencias:
                params['vigencia_0'] = int(vigencias[0])
    if numero_ficha is not None:
        clauses.append('p.numero_ficha = :numero_ficha')
        params['numero_ficha'] = int(numero_ficha)
    if search:
        s = str(search).strip().lower()
        if s:
            clauses.append('LOWER(TRIM(p.denominacion_programa)) LIKE :search')
            params['search'] = f'%{s}%'
    if solo_certificados and str(solo_certificados).strip().lower() not in {'0', 'false', 'no'}:
        clauses.append('(p.certificado IS NOT NULL AND p.certificado <> 0)')

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    sql = (
        get_programas_select_sql() +
        f'{where_sql} '
        'ORDER BY p.fecha_corte DESC, p.numero_ficha ASC, p.id ASC'
    )

    try:
        df = pd.read_sql(text(sql), con=engine, params=params)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error exportando programas: {e}')

    # Exportar todas las columnas de la tabla programas_formacion.
    df_export = df.copy()
    
    # Detectar fecha_corte
    fecha_corte_value = None
    tiene_multiples_fechas = False
    
    if 'fecha_corte' in df_export.columns and not df_export.empty:
        fechas_unicas = df_export['fecha_corte'].dropna().dt.date.unique() if hasattr(df_export['fecha_corte'], 'dt') else pd.Series(df_export['fecha_corte'].dropna()).unique()
        if len(fechas_unicas) > 1:
            tiene_multiples_fechas = True
        elif len(fechas_unicas) == 1:
            fecha_corte_value = str(fechas_unicas[0])
    
    # Remover columna fecha_corte si no hay múltiples fechas
    if 'fecha_corte' in df_export.columns and not tiene_multiples_fechas:
        df_export = df_export.drop('fecha_corte', axis=1)

    # En el frontend se oculta la columna "id"; aqui la mantenemos en el Excel
    # pero la marcamos como oculta para que no aparezca a simple vista.
    original_cols_programas = list(df_export.columns)
    hidden_programas_headers = set()
    if 'id' in original_cols_programas:
        hidden_programas_headers.add('id')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribir datos comenzando desde fila 3 si hay fecha_corte
        if fecha_corte_value:
            df_export.to_excel(writer, index=False, sheet_name='programas', startrow=2)
        else:
            df_export.to_excel(writer, index=False, sheet_name='programas')

        ws = writer.book['programas']
        
        # Agregar fecha_corte en B1
        if fecha_corte_value:
            ws['A1'] = 'Fecha de corte'
            ws['B1'] = fecha_corte_value
            ws['A1'].font = Font(bold=True, size=11)
            ws['B1'].font = Font(size=11)
        
        max_row = ws.max_row
        max_col = ws.max_column

        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = wrap_alignment

        # Verde success de Bootstrap para encabezados
        green_success = "198754"
        header_row = 3 if fecha_corte_value else 1
        for cell in ws[header_row]:
            if cell.value:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color=green_success, end_color=green_success, fill_type="solid")

        min_width = 12
        max_width = 60
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                cell_text = '' if value is None else str(value)
                if len(cell_text) > max_len:
                    max_len = len(cell_text)
            adjusted = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

            # Ocultar columnas que no se muestran en la tabla del frontend.
            header_value = ws.cell(row=header_row, column=col_idx).value
            if header_value in hidden_programas_headers:
                ws.column_dimensions[col_letter].hidden = True
        
        # Ocultar columna fecha_corte si hay múltiples fechas
        if tiene_multiples_fechas and 'fecha_corte' in df.columns:
            fecha_col_idx = list(df.columns).index('fecha_corte')
            col_letter = get_column_letter(fecha_col_idx + 1)
            ws.column_dimensions[col_letter].hidden = True

    output.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'programas_export_{ts}.xlsx'

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.get('/programas/filters')
def get_programas_filters():
    """Devuelve valores distintos para los filtros de programas a nivel global.

    No aplica paginacion ni filtros previos: siempre consulta toda la tabla
    programas_formacion para construir los combos de la UI.
    """
    try:
        with engine.connect() as conn:
            years = [
                int(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT YEAR(fecha_corte) AS y FROM programas_formacion WHERE fecha_corte IS NOT NULL ORDER BY y DESC')
                ).fetchall()
                if r[0] is not None
            ]

            vigencias = [
                int(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT YEAR(fecha_inicio) AS y FROM programas_formacion WHERE fecha_inicio IS NOT NULL ORDER BY y DESC')
                ).fetchall()
                if r[0] is not None
            ]

            municipios = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT ciudad_municipio FROM programas_formacion WHERE ciudad_municipio IS NOT NULL ORDER BY ciudad_municipio ASC')
                ).fetchall()
                if r[0] is not None
            ]

            centros = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT centro_formacion FROM programas_formacion WHERE centro_formacion IS NOT NULL ORDER BY centro_formacion ASC')
                ).fetchall()
                if r[0] is not None
            ]

            niveles = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT nivel_formacion FROM programas_formacion WHERE nivel_formacion IS NOT NULL ORDER BY nivel_formacion ASC')
                ).fetchall()
                if r[0] is not None
            ]

            estrategias = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT estrategia_programa FROM programas_formacion WHERE estrategia_programa IS NOT NULL ORDER BY estrategia_programa ASC')
                ).fetchall()
                if r[0] is not None
            ]

            convenios = [
                str(r[0])
                for r in conn.execute(
                    text('SELECT DISTINCT convenio FROM programas_formacion WHERE convenio IS NOT NULL ORDER BY convenio ASC')
                ).fetchall()
                if r[0] is not None
            ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error obteniendo filtros de programas: {e}')

    return JSONResponse(
        content=jsonable_encoder(
            {
                'years': years,
                'vigencias': vigencias,
                'municipios': municipios,
                'centros': centros,
                'niveles': niveles,
                'estrategias': estrategias,
                'convenios': convenios,
            }
        )
    )


@app.delete('/programas/delete-all')
def delete_programas_all():
    """Elimina todos los registros de programas_formacion."""
    try:
        with engine.begin() as conn:
            result = conn.execute(text('DELETE FROM programas_formacion'))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error eliminando todos los programas: {e}')

    return JSONResponse({'deleted_rows': int(result.rowcount or 0)})


@app.delete('/programas/delete-by-vigencia')
def delete_programas_by_vigencia(vigencia: int):
    """Elimina registros por vigencia (anio de fecha_inicio)."""
    try:
        vig = int(vigencia)
    except Exception:
        raise HTTPException(status_code=400, detail='La vigencia es invalida')

    if vig < 1900 or vig > 2100:
        raise HTTPException(status_code=400, detail='La vigencia debe estar entre 1900 y 2100')

    try:
        with engine.begin() as conn:
            result = conn.execute(
                text('DELETE FROM programas_formacion WHERE YEAR(fecha_inicio) = :vigencia'),
                {'vigencia': vig},
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error eliminando programas por vigencia: {e}')

    return JSONResponse({'vigencia': vig, 'deleted_rows': int(result.rowcount or 0)})


@app.get('/catalogo')
def get_catalogo(
    page: int = 1,
    per_page: int = 50,
    search: Optional[str] = None,
    year: Optional[str] = None,
    nivel: Optional[str] = None,
):
    try:
        page = int(page)
    except Exception:
        page = 1
    try:
        per_page = int(per_page)
    except Exception:
        per_page = 50
    if page < 1:
        page = 1
    if per_page < 1 or per_page > 500:
        per_page = 50

    clauses = []
    params: dict = {}

    if year is not None:
        years = [y.strip() for y in str(year).split(',') if y.strip()]
        if years:
            if len(years) == 1:
                clauses.append('YEAR(fecha_corte) = :year_0')
            else:
                in_keys = []
                for i, val in enumerate(years):
                    key = f'year_{i}'
                    in_keys.append(f':{key}')
                    params[key] = int(val)
                clauses.append('YEAR(fecha_corte) IN (' + ','.join(in_keys) + ')')
            if 'year_0' not in params and years:
                params['year_0'] = int(years[0])

    if search:
        s = str(search).strip().lower()
        if s:
            clauses.append('(LOWER(TRIM(cod_ver)) LIKE :search OR LOWER(TRIM(prf_denominacion)) LIKE :search)')
            params['search'] = f'%{s}%'

    if nivel is not None:
        n = str(nivel).strip().lower()
        if n:
            clauses.append('LOWER(TRIM(nivel_de_formacion)) = :nivel')
            params['nivel'] = n

    where_sql = ''
    if clauses:
        where_sql = ' WHERE ' + ' AND '.join(clauses)

    count_sql = f'SELECT COUNT(*) AS total FROM catalogo{where_sql}'
    try:
        with engine.connect() as conn:
            total = conn.execute(text(count_sql), params).scalar() or 0
            # obtener lista de niveles distintos para poblar el filtro en frontend
            try:
                distinct_sql = "SELECT DISTINCT nivel_de_formacion FROM catalogo WHERE nivel_de_formacion IS NOT NULL AND TRIM(nivel_de_formacion) <> '' ORDER BY nivel_de_formacion"
                res = conn.execute(text(distinct_sql))
                distinct_niveles = [r[0] for r in res.fetchall() if r[0] is not None]
            except Exception:
                distinct_niveles = []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error contando registros de catalogo: {e}')

    offset = (page - 1) * per_page
    sql = (
        'SELECT * FROM catalogo'
        f'{where_sql} '
        'ORDER BY fecha_corte DESC, cod_ver ASC '
        'LIMIT :limit OFFSET :offset'
    )
    params_data = dict(params)
    params_data['limit'] = per_page
    params_data['offset'] = offset

    try:
        df = pd.read_sql(text(sql), con=engine, params=params_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error consultando catalogo: {e}')

    if not df.empty:
        df = df.replace([float('inf'), float('-inf')], pd.NA)
        df = df.where(pd.notna(df), None)
        for col in ['prf_fch_registro', 'fecha_activo_en_ejecucion', 'prf_fecha_resolucion', 'fecha_corte']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                df[col] = df[col].apply(lambda v: v.isoformat() if hasattr(v, 'isoformat') else v)

    items = df.to_dict(orient='records') if not df.empty else []
    
    # Limpiar NaN y inf de los items para JSON serialization
    import math
    for item in items:
        for key in list(item.keys()):
            val = item[key]
            if isinstance(val, float):
                if math.isnan(val) or math.isinf(val):
                    item[key] = None
    latest_fecha = None
    if not df.empty and 'fecha_corte' in df.columns:
        valid = pd.to_datetime(df['fecha_corte'], errors='coerce').dropna()
        if not valid.empty:
            latest_fecha = valid.max().date().isoformat()

    return JSONResponse(
        content=jsonable_encoder(
            {
                'items': items,
                'total': int(total),
                'fecha_corte': latest_fecha,
                'page': page,
                'per_page': per_page,
                'distinct_niveles': distinct_niveles,
            }
        )
    )


@app.post('/catalogo/upload-excel')
async def upload_catalogo_excel(file: UploadFile = File(...), fecha_corte_manual: Optional[date] = Form(None)):
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xml')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls, .xlsx o .xml')

    content = await file.read()

    fecha_corte_file = fecha_corte_manual
    if not fecha_corte_file:
        fecha_corte_file = extract_fecha_corte_from_excel_b1(content)
    if not fecha_corte_file:
        fecha_corte_file = extract_fecha_corte_from_filename(file.filename or '')
    if not fecha_corte_file:
        raise HTTPException(
            status_code=400,
            detail='No se pudo obtener fecha_corte. Envia fecha_corte_manual o usa un archivo con fecha en B1.',
        )

    df = _load_catalogo_dataframe(content, file.filename or '')
    df['fecha_corte'] = fecha_corte_file
    rows_affected = _upsert_catalogo_rows(df)

    return JSONResponse(
        {
            'inserted': int(len(df)),
            'rows_affected': int(rows_affected),
            'fecha_corte': str(fecha_corte_file),
        }
    )


class UpdateRequest(BaseModel):
    cod_fichas: List[int]
    periodo: Optional[int] = None
    oferta: Optional[str] = None
    tipo: Optional[str] = None


class FichaUpdate(BaseModel):
    cod_regional: Optional[int] = None
    regional: Optional[str] = None
    cod_municipio: Optional[int] = None
    municipio: Optional[str] = None
    cod_centro: Optional[int] = None
    centro_formacion: Optional[str] = None
    cod_programa: Optional[int] = None
    denominacion_programa: Optional[str] = None
    cod_ficha: Optional[int] = None
    estado_ficha: Optional[str] = None
    jornada: Optional[str] = None
    nivel_formacion: Optional[str] = None
    cupo: Optional[int] = None
    inscritos_primera_opcion: Optional[int] = None
    inscritos_segunda_opcion: Optional[int] = None
    oferta: Optional[str] = None
    tipo: Optional[str] = None
    perfil_ingreso: Optional[str] = None
    periodo: Optional[int] = None


@app.get('/fichas/{cod_ficha}')
def get_ficha(cod_ficha: int):
    try:
        # Construir la consulta con el id como entero para evitar problemas de parámetros con pymysql
        df = pd.read_sql(f"SELECT * FROM fichas_formacion WHERE cod_ficha = {int(cod_ficha)}", con=engine)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al leer la base de datos: {e}')
    if df.empty:
        raise HTTPException(status_code=404, detail='Ficha no encontrada')
    return JSONResponse(df.iloc[0].to_dict())


@app.put('/fichas/{cod_ficha}')
def update_ficha(cod_ficha: int, payload: FichaUpdate):
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail='No hay campos para actualizar')

    # No permitir cambiar la PK cod_ficha a otro valor desde aquí
    if 'cod_ficha' in data:
        data.pop('cod_ficha')

    updates = {}
    for k, v in data.items():
        if v is None:
            updates[k] = None
        elif k == 'oferta':
            updates['oferta'] = normalize_oferta(v)
        elif k == 'tipo':
            updates['tipo'] = normalize_tipo(v)
        else:
            updates[k] = v

    set_parts = []
    params = {}
    for i, (k, v) in enumerate(updates.items()):
        param_name = f'val_{i}'
        set_parts.append(f"{k} = :{param_name}")
        params[param_name] = v

    params['id'] = cod_ficha
    sql = text(f"UPDATE fichas_formacion SET {', '.join(set_parts)} WHERE cod_ficha = :id")

    try:
        with engine.begin() as conn:
            result = conn.execute(sql, params)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al actualizar la ficha: {e}')

    return JSONResponse({'updated_rows': result.rowcount})


@app.post('/fichas/update')
def update_fichas(req: UpdateRequest):
    if not req.cod_fichas:
        raise HTTPException(status_code=400, detail='Se requiere al menos un cod_ficha')

    updates = {}
    if req.periodo is not None:
        updates['periodo'] = int(req.periodo)
    if req.oferta is not None:
        updates['oferta'] = normalize_oferta(req.oferta)
    if req.tipo is not None:
        updates['tipo'] = normalize_tipo(req.tipo)

    if not updates:
        raise HTTPException(status_code=400, detail='No hay campos para actualizar')

    set_parts = []
    params = {}
    for i, (k, v) in enumerate(updates.items()):
        param_name = f'val_{i}'
        set_parts.append(f"{k} = :{param_name}")
        params[param_name] = v

    params['ids'] = req.cod_fichas
    sql = text(f"UPDATE fichas_formacion SET {', '.join(set_parts)} WHERE cod_ficha IN :ids").bindparams(bindparam('ids', expanding=True))

    try:
        with engine.begin() as conn:
            result = conn.execute(sql, params)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al actualizar registros: {e}')

    return JSONResponse({'updated_rows': result.rowcount})


# ============================================================================
# FUNCIONES Y ENDPOINTS PARA CONSOLIDADO COLEGIOS
# ============================================================================

def _read_excel_consolidado_colegios(content: bytes) -> pd.DataFrame:
    """Lee un archivo Excel de consolidado colegios"""
    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, dtype=str)
        return df
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f'Error al leer el archivo Excel: {str(e)}'
        )


def _normalize_col_name_consolidado(col_name: str) -> str:
    """Normaliza nombres de columna para búsqueda flexible"""
    text = str(col_name).lower().strip()
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    text = text.replace(' ', '_').replace('-', '_')
    return text


async def _process_consolidado_colegios(df: pd.DataFrame):
    """Procesa el DataFrame de consolidado colegios e inserta en la tabla"""
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    # Normalizar nombres de columnas
    df.columns = [_normalize_col_name_consolidado(col) for col in df.columns]

    # Mapear encabezados del Excel a nombres de columnas de la tabla
    column_mapping = {
        'nombre_real_de_la_institucion': 'nombre_real_institucion',
        'nombre_real_institucion': 'nombre_real_institucion',
        'nombres_registrados_en_sofia_plus': 'nombres_sofia_plus',
        'nombres_sofia_plus': 'nombres_sofia_plus',
        'nombres_registrados_sofia': 'nombres_sofia_plus',
        'municipio': 'municipio',
        'clasificacion': 'clasificacion',
        'clasificaci_n': 'clasificacion',
    }

    df_mapped = df.copy()
    for original, mapped in column_mapping.items():
        if original in df.columns and mapped not in df.columns:
            df_mapped[mapped] = df[original]

    # Seleccionar solo las columnas mapeadas
    required_cols = ['nombre_real_institucion', 'nombres_sofia_plus', 'municipio', 'clasificacion']
    for col in required_cols:
        if col not in df_mapped.columns:
            df_mapped[col] = None

    df_mapped = df_mapped[required_cols].copy()

    # Limpiar datos
    def clean_text(val):
        if pd.isna(val):
            return None
        text = str(val).strip()
        return text if text else None

    for col in df_mapped.columns:
        df_mapped[col] = df_mapped[col].apply(clean_text)

    # Convertir NaN a None
    df_mapped = df_mapped.where(pd.notna(df_mapped), None)

    # Eliminar filas completamente vacías
    df_mapped = df_mapped.dropna(how='all').reset_index(drop=True)
    if df_mapped.empty:
        raise HTTPException(status_code=400, detail='No hay registros válidos después de validar')

    try:
        with engine.connect() as conn:
            inserted_count = 0
            for idx, row in df_mapped.iterrows():
                values = {}
                for col in df_mapped.columns:
                    val = row[col]
                    if pd.isna(val):
                        values[col] = None
                    else:
                        values[col] = str(val).strip() if val else None

                cols = list(values.keys())
                col_names = ', '.join([f'`{c}`' for c in cols])
                placeholders = ', '.join([f':{c}' for c in cols])

                insert_sql = f"""
                INSERT INTO consolidado_colegios ({col_names})
                VALUES ({placeholders})
                """

                try:
                    conn.execute(text(insert_sql), values)
                    inserted_count += 1
                except Exception as row_error:
                    print(f"Error insertando fila {idx}: {row_error}")
                    print(f"Valores: {values}")

            conn.commit()

        return {
            'status': 'success',
            'message': f'Se cargaron {inserted_count} registros correctamente en consolidado_colegios',
            'inserted': inserted_count
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al guardar en base de datos: {str(e)}'
        )


@app.post('/consolidado-colegios/upload-excel')
async def upload_consolidado_colegios_excel(file: UploadFile = File(...)):
    """Sube un archivo Excel de Consolidado Colegios."""
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls o .xlsx')

    content = await file.read()

    try:
        df = _read_excel_consolidado_colegios(content)
        result = await _process_consolidado_colegios(df)
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al procesar archivo: {str(e)}')


@app.get('/consolidado-colegios/data')
async def get_consolidado_colegios_data():
    """Obtiene los datos cargados de Consolidado Colegios."""
    try:
        with engine.connect() as conn:
            # Verificar si la tabla existe
            check_table_sql = """
            SELECT COUNT(*) as count
            FROM information_schema.TABLES 
            WHERE TABLE_SCHEMA = DATABASE() 
            AND TABLE_NAME = 'consolidado_colegios'
            """
            table_exists = conn.execute(text(check_table_sql)).fetchone()

            if not table_exists or table_exists[0] == 0:
                return JSONResponse({'items': [], 'total': 0, 'message': 'Tabla no existe aún'})

            # Obtener TODOS los datos de la tabla
            result = conn.execute(text("""
                SELECT 
                    id,
                    nombre_real_institucion,
                    nombres_sofia_plus,
                    municipio,
                    clasificacion,
                    fecha_registro
                FROM consolidado_colegios
                ORDER BY fecha_registro DESC
                LIMIT 10000
            """))
            rows = result.fetchall()
            data = []
            for row in rows:
                row_dict = dict(row._mapping)
                # Convertir fechas a ISO format string
                for key in list(row_dict.keys()):
                    val = row_dict[key]
                    if val is not None and hasattr(val, 'isoformat'):
                        row_dict[key] = val.isoformat()
                data.append(row_dict)

            return JSONResponse(content=jsonable_encoder({'items': data, 'total': len(data)}))
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse({'items': [], 'error': str(e), 'detail': error_detail}, status_code=500)


# ============================================================================
# MINI MÓDULO: AGREGAR PROGRAMAS EN CATÁLOGO
# ============================================================================

def _read_excel_agregar_programas(content: bytes) -> pd.DataFrame:
    """Lee un archivo Excel con códigos prf_codigo"""
    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, dtype=str)
        return df
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f'Error al leer el archivo Excel: {str(e)}'
        )


async def _process_agregar_programas(df: pd.DataFrame, tipo_programa: str):
    """Procesa el DataFrame de códigos y actualiza catalogo"""
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    # Validar que tipo_programa sea válido
    valid_tipos = ['Construcción FEC', 'Transformación FEC', 'FEP']
    if tipo_programa not in valid_tipos:
        raise HTTPException(
            status_code=400,
            detail=f'Tipo de programa inválido. Debe ser: {", ".join(valid_tipos)}'
        )

    # Normalizar nombres de columnas
    df.columns = [str(col).lower().strip().replace(' ', '_').replace('-', '_') for col in df.columns]

    # Buscar columna con código (puede ser: prf_codigo, codigo, codigo_programa, etc.)
    codigo_col = None
    for col in df.columns:
        if 'codigo' in col or 'prf' in col or 'code' in col:
            codigo_col = col
            break

    if not codigo_col:
        raise HTTPException(
            status_code=400,
            detail='No se encontró columna de código. Esperaba una columna con "código" o "prf_codigo" en el nombre'
        )

    # Extraer códigos únicos
    codigos = df[codigo_col].dropna().astype(str).str.strip()
    codigos = codigos[codigos != ''].unique().tolist()

    if not codigos:
        raise HTTPException(status_code=400, detail='No hay códigos válidos en el archivo Excel')

    try:
        with engine.connect() as conn:
            updated_count = 0
            not_found_count = 0

            for codigo in codigos:
                try:
                    codigo_int = int(float(codigo))
                except (ValueError, TypeError):
                    not_found_count += 1
                    continue

                # Buscar en catalogo
                check_sql = text("""
                    SELECT prf_codigo FROM catalogo 
                    WHERE prf_codigo = :codigo
                    LIMIT 1
                """)

                result = conn.execute(check_sql, {'codigo': codigo_int}).fetchone()

                if result:
                    # Actualizar el registro
                    update_sql = text("""
                        UPDATE catalogo 
                        SET tipo_programa = :tipo_programa
                        WHERE prf_codigo = :codigo
                    """)

                    conn.execute(update_sql, {
                        'tipo_programa': tipo_programa,
                        'codigo': codigo_int
                    })
                    updated_count += 1
                else:
                    not_found_count += 1

            conn.commit()

        return {
            'status': 'success',
            'message': f'Procesados {len(codigos)} códigos',
            'updated': updated_count,
            'not_found': not_found_count,
            'tipo_programa': tipo_programa
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al guardar en base de datos: {str(e)}'
        )


@app.post('/catalogo/agregar-programas')
async def agregar_programas_catalogo(
    file: UploadFile = File(...),
    tipo_programa: str = Form(...)
):
    """Sube un Excel con códigos y marca los programas con el tipo seleccionado."""
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls o .xlsx')

    content = await file.read()

    try:
        df = _read_excel_agregar_programas(content)
        result = await _process_agregar_programas(df, tipo_programa)
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al procesar archivo: {str(e)}')


@app.get('/catalogo/programas-por-tipo')
async def get_programas_por_tipo(tipo_programa: Optional[str] = None):
    """Obtiene programas filtrados por tipo_programa."""
    try:
        with engine.connect() as conn:
            if tipo_programa:
                result = conn.execute(text("""
                    SELECT 
                        prf_codigo,
                        cod_ver,
                        prf_denominacion,
                        nivel_de_formacion,
                        tipo_programa
                    FROM catalogo
                    WHERE tipo_programa = :tipo_programa
                    ORDER BY prf_codigo
                    LIMIT 5000
                """), {'tipo_programa': tipo_programa})
            else:
                result = conn.execute(text("""
                    SELECT 
                        prf_codigo,
                        cod_ver,
                        prf_denominacion,
                        nivel_de_formacion,
                        tipo_programa
                    FROM catalogo
                    WHERE tipo_programa IS NOT NULL
                    ORDER BY tipo_programa, prf_codigo
                    LIMIT 5000
                """))

            rows = result.fetchall()
            data = [dict(row._mapping) for row in rows]
            
            return JSONResponse(content=jsonable_encoder({
                'items': data,
                'total': len(data),
                'tipo_programa': tipo_programa or 'Todos'
            }))

    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse(
            {'items': [], 'error': str(e), 'detail': error_detail},
            status_code=500
        )


# ===== ENDPOINTS DE EXPORTACIÓN A EXCEL =====

def _export_table_to_excel(table_name: str, output_filename: str) -> StreamingResponse:
    """
    Función genérica para exportar cualquier tabla a Excel
    - Verde oscuro en encabezados
    - Remueve columna fecha_corte de los datos
    - Pone fecha de corte en B1 (si existe)
    - Oculta columna fecha_corte si hay múltiples fechas
    """
    try:
        sql = f'SELECT * FROM {table_name}'
        
        df = pd.read_sql(text(sql), con=engine)
        
        if df.empty:
            df = pd.DataFrame()
        
        # Detectar si existe columna fecha_corte y obtener información
        fecha_corte_value = None
        tiene_multiples_fechas = False
        
        if 'fecha_corte' in df.columns and not df.empty:
            fechas_unicas = df['fecha_corte'].dropna().unique()
            if len(fechas_unicas) > 1:
                tiene_multiples_fechas = True
            elif len(fechas_unicas) == 1:
                fecha_corte_value = str(fechas_unicas[0])
        
        # Remover columna fecha_corte de los datos si no hay múltiples fechas
        if 'fecha_corte' in df.columns and not tiene_multiples_fechas:
            df = df.drop('fecha_corte', axis=1)
        elif 'fecha_corte' in df.columns and tiene_multiples_fechas:
            # Mantener la columna pero la ocultaremos después
            pass
        
        # Convertir columnas datetime a string para mejor manejo en Excel
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escribir datos comenzando desde fila 3 (para dejar espacio para fecha de corte)
            if fecha_corte_value:
                df.to_excel(writer, sheet_name='Datos', index=False, startrow=2)
            else:
                df.to_excel(writer, sheet_name='Datos', index=False)
            
            # Dar formato a la hoja
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # Agregar fecha de corte en B1 si existe
            if fecha_corte_value:
                worksheet['A1'] = 'Fecha de corte'
                worksheet['B1'] = fecha_corte_value
                worksheet['A1'].font = Font(bold=True, size=11)
                worksheet['B1'].font = Font(size=11)
            
            # Verde success de Bootstrap (color institucional)
            green_success = "198754"  # Bootstrap success color
            
            # Formatear encabezados (fila 3 si hay fecha_corte, fila 1 si no)
            header_row = 3 if fecha_corte_value else 1
            for cell in worksheet[header_row]:
                if cell.value:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color=green_success, end_color=green_success, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Ocultar columna fecha_corte si hay múltiples fechas
            if tiene_multiples_fechas and 'fecha_corte' in df.columns:
                # Encontrar el índice de la columna fecha_corte
                fecha_corte_col_index = list(df.columns).index('fecha_corte')
                col_letter = get_column_letter(fecha_corte_col_index + 1)
                worksheet.column_dimensions[col_letter].hidden = True
        
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{output_filename}"'}
        )
    
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=f'Error exportando {table_name}: {str(e)}\n{traceback.format_exc()}')


@app.get('/catalogo/exportar-excel')
def export_catalogo():
    """Exportar tabla catalogo a Excel"""
    return _export_table_to_excel('catalogo', 'catalogo.xlsx')


@app.get('/registro-calificado/exportar-excel')
def export_registro_calificado():
    """Exportar tabla registro_calificado_presencial a Excel"""
    return _export_table_to_excel('registro_calificado_presencial', 'registro_calificado_presencial.xlsx')


@app.get('/oferta/exportar-excel')
def export_oferta():
    """Exportar tabla OFERTA_seguimiento_metas a Excel"""
    return _export_table_to_excel('OFERTA_seguimiento_metas', 'oferta_seguimiento_metas.xlsx')


@app.get('/consolidado-colegios/exportar-excel')
def export_consolidado_colegios():
    """Exportar tabla consolidado_colegios a Excel"""
    return _export_table_to_excel('consolidado_colegios', 'consolidado_colegios.xlsx')


# ============================================================================
# MINI MÓDULO: PE_04 SEGUIMIENTO DE METAS
# ============================================================================

def _read_excel_pe04(content: bytes) -> pd.DataFrame:
    """Lee un archivo Excel PE_04 de programas de formación con encabezados en fila 2"""
    try:
        # header=1 significa que los encabezados están en la fila 2 (índice 1, basado en 0)
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, header=1, dtype=str)
        return df
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f'Error al leer el archivo Excel: {str(e)}'
        )


def _normalize_col_name_pe04(col_name: str) -> str:
    """Normaliza nombres de columna para búsqueda flexible"""
    text = str(col_name).lower().strip()
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    text = text.replace(' ', '_').replace('-', '_')
    return text


def _clasificar_programa_especial(nombre_programa_especial: str, nombre_convenio: str) -> str:
    """
    Clasifica los programas según la fórmula compleja del Excel PE_04
    Retorna: SENATEC, ACME, SER CAMPESENA, SER, BILINGUISMO, CAMPESENA, ECONOMIA POPULAR, 
             CAMPESENA RADIAL, FIC, o NA
    """
    if not nombre_programa_especial:
        nombre_programa_especial = ""
    if not nombre_convenio:
        nombre_convenio = ""
    
    nombre_programa_especial = str(nombre_programa_especial).strip().upper()
    nombre_convenio = str(nombre_convenio).strip().upper()
    
    # 1. SENATEC
    if nombre_programa_especial == "SENATEC":
        return "SENATEC"
    
    # 2. ACME
    if nombre_programa_especial in ["INTEGRACIÓN CON LA EDUCACIÓN MEDIA ACADÉMICA", 
                                     "INTEGRACIÓN CON LA EDUCACIÓN MEDIA TÉCNICA"]:
        return "ACME"
    
    # 3. SER CAMPESENA
    acuerdo_campesena = "ACUERDO NO. 0003 DE 2023  POR EL CUAL SE CREA LA ESTRATEGIA CAMPE-SENA" in nombre_convenio or \
                        "ACUERDO NO. 0003 DE 2023 POR EL CUAL SE CREA LA ESTRATEGIA CAMPE-SENA" in nombre_convenio
    
    if nombre_programa_especial == "CAMPESENA- SER":
        return "SER CAMPESENA"
    
    if nombre_programa_especial == "SER" and acuerdo_campesena:
        return "SER CAMPESENA"
    
    # 4. SER
    if nombre_programa_especial == "SER":
        return "SER"
    
    # 5. BILINGUISMO
    if nombre_programa_especial == "PROGRAMA DE BILINGUISMO":
        return "BILINGUISMO"
    
    # 6. CAMPESENA
    campesena_keywords = ["CAMPESENA- AULA MÓVIL", "CAMPESENA", "FORMACIÓN CONTINUA ESPECIAL CAMPESINA"]
    if acuerdo_campesena or any(kw in nombre_programa_especial for kw in campesena_keywords):
        return "CAMPESENA"
    
    # 7. ECONOMIA POPULAR
    economia_keywords = ["FULL POPULAR  FORMACIÓN CONTINUA ESPECIAL POPULAR", 
                        "ECONOMIA POPULAR- AULA MÓVIL", "FULL POPULAR"]
    economia_convenio = "ECONOMÍA POPULAR - PND - SENA" in nombre_convenio
    if any(kw in nombre_programa_especial for kw in economia_keywords) or economia_convenio:
        return "ECONOMIA POPULAR"
    
    # 8. CAMPESENA RADIAL
    if nombre_programa_especial == "CAMPESENA RADIAL":
        return "CAMPESENA RADIAL"
    
    # 9. FIC
    if "CURSOS FIC" in nombre_convenio:
        return "FIC"
    
    # 10. NA (default)
    return "NA"


async def _process_pe04(df: pd.DataFrame):
    """Procesa el DataFrame PE_04 e inserta en programas_formacion_seguimiento_pe04"""
    if df.empty:
        raise HTTPException(status_code=400, detail='El Excel no contiene filas')

    # Mapear directamente los encabezados reales del Excel a nombres de columnas de la tabla
    # Esto es más robusto que normalizar ya que los encabezados pueden variar
    column_mapping_excel_to_db = {
        'NOMBRE_CENTRO': 'centro_formacion',
        'IDENTIFICADOR_FICHA': 'numero_ficha',
        'NOMBRE_MUNICIPIO_CURSO': 'ciudad_municipio',
        'FECHA_INICIO_FICHA': 'fecha_inicio',
        'FECHA_TERMINACION_FICHA': 'fecha_fin',
        'NIVEL_FORMACION': 'nivel_formacion',
        'NOMBRE_PROGRAMA_FORMACION': 'denominacion_programa',
        'NOMBRE_SECTOR_PROGRAMA': 'estrategia_programa',
        'NOMBRE_CONVENIO': 'convenio',
        'NUMERO_CURSOS': 'cupos',
        'TOTAL_APRENDICES_ACTIVOS': 'aprendices_activos',
        'TIPO_DE_FORMACION': 'tipo_formacion',
        'MODALIDAD_FORMACION': 'modalidad_formacion',
        'ESTADO_CURSO': 'estado_curso',
        'NOMBRE_EMPRESA': 'nombre_empresa',
    }

    # Crear nuevo DataFrame con solo las columnas que necesitamos
    df_mapped = pd.DataFrame()
    
    for excel_col, db_col in column_mapping_excel_to_db.items():
        if excel_col in df.columns:
            df_mapped[db_col] = df[excel_col]
        else:
            # Si la columna no existe, crear columna vacía
            df_mapped[db_col] = None
    
    # Agregar certificado y fecha_corte como columnas vacías (no existen en el Excel)
    df_mapped['certificado'] = None
    df_mapped['fecha_corte'] = None
    
    # Inicializar columnas nuevas que se calcularán
    df_mapped['aprendices_matriculados'] = None
    df_mapped['fecha_inicio_etapa_productiva'] = None
    df_mapped['vigencia_aprendices'] = None
    
    # Agregar clasificación de programa especial basada en los datos del Excel
    # Obtener NOMBRE_PROGRAMA_ESPECIAL del DataFrame original si existe
    nombre_programa_especial_col = [col for col in df.columns if 'NOMBRE_PROGRAMA_ESPECIAL' in col]
    
    if nombre_programa_especial_col:
        programa_especial_data = df[nombre_programa_especial_col[0]]
    else:
        programa_especial_data = [None] * len(df)
    
    # Aplicar clasificación a cada fila
    df_mapped['clasificacion_programa_especial'] = [
        _clasificar_programa_especial(
            programa_especial_data.iloc[i] if i < len(programa_especial_data) else None,
            df_mapped['convenio'].iloc[i] if i < len(df_mapped) else None
        )
        for i in range(len(df_mapped))
    ]
    
    # Calcular fecha_inicio_etapa_productiva basada en nivel_formacion y fecha_fin
    # Y asignar vigencia_aprendices basada en fecha_inicio
    from datetime import datetime as dt, timedelta
    from dateutil.relativedelta import relativedelta
    
    def calcular_inicio_etapa_productiva(nivel, fecha_fin):
        """Calcula la fecha de inicio de etapa productiva según el nivel de formación"""
        if pd.isna(fecha_fin) or fecha_fin is None:
            return None
        
        try:
            # Convertir a date si es necesario
            if isinstance(fecha_fin, str):
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                    try:
                        fecha_fin = dt.strptime(fecha_fin.strip(), fmt).date()
                        break
                    except:
                        pass
            elif isinstance(fecha_fin, dt):
                fecha_fin = fecha_fin.date()
            
            if not isinstance(fecha_fin, (dt.date() if hasattr(dt, 'date') else date.__class__)):
                return None
            
            nivel_str = str(nivel).strip().upper() if nivel else ""
            
            # Aplicar lógica según nivel
            if nivel_str in ['TECNÓLOGO', 'TECNICO', 'TÉCNICO']:
                # 6 meses atrás
                fecha_inicio = fecha_fin - relativedelta(months=6)
            elif nivel_str in ['OPERARIO', 'AUXILIAR']:
                # 3 meses atrás
                fecha_inicio = fecha_fin - relativedelta(months=3)
            else:
                # Para otros niveles: "No aplica" se manejará después
                return None
            
            return fecha_inicio
        except Exception as e:
            print(f"Error calculando etapa productiva: {e}")
            return None
    
    # Aplicar cálculo de etapa productiva
    df_mapped['fecha_inicio_etapa_productiva'] = [
        calcular_inicio_etapa_productiva(
            df_mapped['nivel_formacion'].iloc[i],
            df_mapped['fecha_fin'].iloc[i]
        ) if i < len(df_mapped) else None
        for i in range(len(df_mapped))
    ]
    
    # Asignar vigencia basada en fecha_inicio
    def extraer_vigencia(fecha_inicio):
        """Extrae el año de la fecha de inicio como vigencia"""
        try:
            if pd.isna(fecha_inicio) or fecha_inicio is None:
                return None
            
            if isinstance(fecha_inicio, str):
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                    try:
                        fecha = dt.strptime(fecha_inicio.strip(), fmt).date()
                        return fecha.year
                    except:
                        pass
            elif isinstance(fecha_inicio, dt):
                return fecha_inicio.year
            elif hasattr(fecha_inicio, 'year'):
                return fecha_inicio.year
            
            return None
        except:
            return None
    
    df_mapped['vigencia_aprendices'] = [
        extraer_vigencia(df_mapped['fecha_inicio'].iloc[i]) if i < len(df_mapped) else None
        for i in range(len(df_mapped))
    ]
    
    # Por ahora, aprendices_matriculados será igual a aprendices_activos
    # Se actualizará con lógica más compleja después en el procesamiento de la base de datos
    df_mapped['aprendices_matriculados'] = df_mapped['aprendices_activos'].copy()

    # Limpiar datos
    def clean_value(val, col_type='text'):
        if pd.isna(val) or val is None or val == '':
            return None
        if col_type in ['date']:
            try:
                from datetime import datetime
                if isinstance(val, datetime):
                    return val.date()
                elif isinstance(val, str):
                    # Intentar varios formatos de fecha
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                        try:
                            return datetime.strptime(val.strip(), fmt).date()
                        except:
                            pass
                    return None
                else:
                    # Si es un número (Excel almacena fechas como números)
                    try:
                        from datetime import datetime, timedelta
                        base_date = datetime(1899, 12, 30)  # Excel epoch
                        return (base_date + timedelta(days=float(val))).date()
                    except:
                        return None
            except:
                return None
        elif col_type in ['int']:
            try:
                return int(float(str(val)))
            except:
                return None
        else:
            text = str(val).strip()
            return text if text else None

    # Aplicar limpieza específica por tipo de columna
    date_cols = ['fecha_inicio', 'fecha_fin', 'fecha_corte', 'fecha_inicio_etapa_productiva']
    int_cols = ['numero_ficha', 'cupos', 'aprendices_activos', 'aprendices_matriculados', 'vigencia_aprendices']
    
    for col in df_mapped.columns:
        if col in date_cols:
            df_mapped[col] = df_mapped[col].apply(lambda x: clean_value(x, 'date'))
        elif col in int_cols:
            df_mapped[col] = df_mapped[col].apply(lambda x: clean_value(x, 'int'))
        else:
            df_mapped[col] = df_mapped[col].apply(lambda x: clean_value(x, 'text'))

    # Convertir NaN a None
    df_mapped = df_mapped.where(pd.notna(df_mapped), None)

    # Eliminar filas completamente vacías
    df_mapped = df_mapped.dropna(how='all').reset_index(drop=True)
    if df_mapped.empty:
        raise HTTPException(status_code=400, detail='No hay registros válidos después de validar')

    try:
        with engine.connect() as conn:
            # Primero, limpiar la tabla (solo lectura, así que cada carga reemplaza)
            conn.execute(text('DELETE FROM programas_formacion_seguimiento_pe04'))
            conn.commit()

            inserted_count = 0
            for idx, row in df_mapped.iterrows():
                values = {}
                for col in df_mapped.columns:
                    val = row[col]
                    if pd.isna(val):
                        values[col] = None
                    elif isinstance(val, (int, float)):
                        values[col] = val
                    elif col in date_cols:
                        values[col] = val
                    else:
                        values[col] = str(val).strip() if val else None

                cols = list(values.keys())
                col_names = ', '.join([f'`{c}`' for c in cols])
                placeholders = ', '.join([f':{c}' for c in cols])

                insert_sql = f"""
                INSERT INTO programas_formacion_seguimiento_pe04 ({col_names})
                VALUES ({placeholders})
                """

                try:
                    conn.execute(text(insert_sql), values)
                    inserted_count += 1
                except Exception as row_error:
                    print(f"Error insertando fila {idx}: {row_error}")
                    print(f"Valores: {values}")

            conn.commit()

        return {
            'status': 'success',
            'message': f'Se cargaron {inserted_count} registros PE_04 correctamente',
            'inserted': inserted_count
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al guardar en base de datos: {str(e)}'
        )


@app.post('/pe04-seguimiento/upload-excel')
async def upload_pe04_excel(file: UploadFile = File(...)):
    """Sube un archivo Excel PE_04 (Seguimiento de Metas)."""
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='El archivo debe ser .xls o .xlsx')

    content = await file.read()

    try:
        df = _read_excel_pe04(content)
        result = await _process_pe04(df)
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Error al procesar archivo: {str(e)}')


@app.get('/pe04-seguimiento/data')
async def get_pe04_data():
    """Obtiene los datos cargados de PE_04 (Seguimiento de Metas)."""
    try:
        with engine.connect() as conn:
            # Verificar si la tabla existe
            check_table_sql = """
            SELECT COUNT(*) as count
            FROM information_schema.TABLES 
            WHERE TABLE_SCHEMA = DATABASE() 
            AND TABLE_NAME = 'programas_formacion_seguimiento_pe04'
            """
            table_exists = conn.execute(text(check_table_sql)).fetchone()

            if not table_exists or table_exists[0] == 0:
                return JSONResponse({'items': [], 'total': 0, 'message': 'Tabla no existe aún'})

            # Obtener TODOS los datos de la tabla
            result = conn.execute(text("""
                SELECT 
                    id,
                    centro_formacion,
                    numero_ficha,
                    ciudad_municipio,
                    fecha_inicio,
                    fecha_fin,
                    nivel_formacion,
                    denominacion_programa,
                    estrategia_programa,
                    convenio,
                    cupos,
                    aprendices_activos,
                    aprendices_matriculados,
                    certificado,
                    tipo_formacion,
                    modalidad_formacion,
                    estado_curso,
                    fecha_corte,
                    clasificacion_programa_especial,
                    nombre_empresa,
                    fecha_inicio_etapa_productiva,
                    vigencia_aprendices,
                    fecha_carga
                FROM programas_formacion_seguimiento_pe04
                ORDER BY fecha_carga DESC
                LIMIT 10000
            """))
            rows = result.fetchall()
            data = []
            for row in rows:
                row_dict = dict(row._mapping)
                # Convertir fechas a ISO format string
                for key in list(row_dict.keys()):
                    val = row_dict[key]
                    if val is not None and hasattr(val, 'isoformat'):
                        row_dict[key] = val.isoformat()
                data.append(row_dict)

            return JSONResponse(content=jsonable_encoder({'items': data, 'total': len(data)}))
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse({'items': [], 'error': str(e), 'detail': error_detail}, status_code=500)


@app.get('/pe04-seguimiento/exportar-excel')
def export_pe04():
    """Exportar tabla programas_formacion_seguimiento_pe04 a Excel"""
    return _export_table_to_excel('programas_formacion_seguimiento_pe04', 'pe04_seguimiento.xlsx')


@app.get('/pe04-seguimiento/resumen-modalidades')
async def get_pe04_resumen_modalidades(centro: str = None):
    """Resumen de PRESENCIAL, VIRTUAL, A DISTANCIA (excluye ECONOMIA POPULAR, FIC, CAMPESENA):
    - NIVEL_FORMACION = 'TECNÓLOGO'
    - Fichas que pasan 2026 (fecha_fin > 2025-12-31)
    - Parámetro opcional: ?centro=nombre_centro
    """
    try:
        with engine.connect() as conn:
            query = """
            SELECT 
                centro_formacion,
                modalidad_formacion,
                clasificacion_programa_especial,
                COUNT(DISTINCT numero_ficha) as total_fichas,
                COALESCE(SUM(aprendices_activos), 0) as total_aprendices
            FROM programas_formacion_seguimiento_pe04
            WHERE TRIM(UPPER(nivel_formacion)) = 'TECNOLOGO'
              AND fecha_fin > '2025-12-31'
              AND clasificacion_programa_especial NOT IN ('ECONOMIA POPULAR', 'FIC', 'CAMPESENA')
            """
            
            if centro:
                query += f" AND centro_formacion = '{centro}'"
            
            query += """
            GROUP BY centro_formacion, modalidad_formacion, clasificacion_programa_especial
            ORDER BY centro_formacion ASC, modalidad_formacion ASC, total_aprendices DESC
            """
            
            result = conn.execute(text(query))
            rows = result.fetchall()
            
            data = []
            total_general = 0
            for row in rows:
                row_dict = dict(row._mapping)
                total_general += row_dict.get('total_aprendices', 0)
                data.append(row_dict)
            
            # Obtener lista de centros disponibles
            centros_result = conn.execute(text("""
                SELECT DISTINCT centro_formacion
                FROM programas_formacion_seguimiento_pe04
                WHERE TRIM(UPPER(nivel_formacion)) = 'TECNOLOGO'
                  AND fecha_fin > '2025-12-31'
                  AND clasificacion_programa_especial NOT IN ('ECONOMIA POPULAR', 'FIC', 'CAMPESENA')
                ORDER BY centro_formacion ASC
            """))
            centros = [row[0] for row in centros_result.fetchall()]
            
            return JSONResponse(content=jsonable_encoder({
                'items': data, 
                'total': len(data),
                'total_aprendices': total_general,
                'centros_disponibles': centros,
                'centro_seleccionado': centro
            }))
    
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse({'items': [], 'error': str(e), 'detail': error_detail}, status_code=500)


@app.get('/pe04-seguimiento/resumen-especiales')
async def get_pe04_resumen_especiales(centro: str = None):
    """Resumen de ECONOMIA POPULAR, FIC, CAMPESENA (sin agrupar por modalidad):
    - NIVEL_FORMACION = 'TECNÓLOGO'
    - Fichas que pasan 2026 (fecha_fin > 2025-12-31)
    - Parámetro opcional: ?centro=nombre_centro
    """
    try:
        with engine.connect() as conn:
            query = """
            SELECT 
                centro_formacion,
                clasificacion_programa_especial,
                COUNT(DISTINCT numero_ficha) as total_fichas,
                COALESCE(SUM(aprendices_activos), 0) as total_aprendices
            FROM programas_formacion_seguimiento_pe04
            WHERE TRIM(UPPER(nivel_formacion)) = 'TECNOLOGO'
              AND fecha_fin > '2025-12-31'
              AND clasificacion_programa_especial IN ('ECONOMIA POPULAR', 'FIC', 'CAMPESENA')
            """
            
            if centro:
                query += f" AND centro_formacion = '{centro}'"
            
            query += """
            GROUP BY centro_formacion, clasificacion_programa_especial
            ORDER BY centro_formacion ASC, total_aprendices DESC
            """
            
            result = conn.execute(text(query))
            rows = result.fetchall()
            
            data = []
            total_general = 0
            for row in rows:
                row_dict = dict(row._mapping)
                total_general += row_dict.get('total_aprendices', 0)
                data.append(row_dict)
            
            # Obtener lista de centros disponibles
            centros_result = conn.execute(text("""
                SELECT DISTINCT centro_formacion
                FROM programas_formacion_seguimiento_pe04
                WHERE TRIM(UPPER(nivel_formacion)) = 'TECNOLOGO'
                  AND fecha_fin > '2025-12-31'
                  AND clasificacion_programa_especial IN ('ECONOMIA POPULAR', 'FIC', 'CAMPESENA')
                ORDER BY centro_formacion ASC
            """))
            centros = [row[0] for row in centros_result.fetchall()]
            
            return JSONResponse(content=jsonable_encoder({
                'items': data, 
                'total': len(data),
                'total_aprendices': total_general,
                'centros_disponibles': centros,
                'centro_seleccionado': centro
            }))
    
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse({'items': [], 'error': str(e), 'detail': error_detail}, status_code=500)


@app.get('/pe04-seguimiento/debug-filtros')
async def debug_filtros():
    """Debug endpoint para ver qué valores existen en la tabla PE_04"""
    try:
        with engine.connect() as conn:
            debug_data = {
                'total_registros': 0,
                'niveles_formacion': [],
                'tipos_formacion': [],
                'fechas_rango': {},
                'fichas_pasan_2026': 0,
                'fichas_tecnologos': 0,
                'fichas_presencial': 0,
                'fichas_todas_condiciones': 0,
                'ejemplo_registros_todas_condiciones': []
            }
            
            # Total registros
            result = conn.execute(text("SELECT COUNT(*) as cnt FROM programas_formacion_seguimiento_pe04"))
            debug_data['total_registros'] = result.fetchone()[0]
            
            # Niveles únicos
            result = conn.execute(text("SELECT DISTINCT nivel_formacion FROM programas_formacion_seguimiento_pe04 WHERE nivel_formacion IS NOT NULL"))
            debug_data['niveles_formacion'] = [row[0] for row in result.fetchall()]
            
            # Tipos únicos
            result = conn.execute(text("SELECT DISTINCT tipo_formacion FROM programas_formacion_seguimiento_pe04 WHERE tipo_formacion IS NOT NULL"))
            debug_data['tipos_formacion'] = [row[0] for row in result.fetchall()]
            
            # Rango de fechas
            result = conn.execute(text("SELECT MIN(fecha_fin) as min_fecha, MAX(fecha_fin) as max_fecha FROM programas_formacion_seguimiento_pe04"))
            row = result.fetchone()
            if row[0]:
                debug_data['fechas_rango'] = {
                    'min': row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0]),
                    'max': row[1].isoformat() if hasattr(row[1], 'isoformat') else str(row[1])
                }
            
            # Fichas pasan 2026
            result = conn.execute(text("SELECT COUNT(*) as cnt FROM programas_formacion_seguimiento_pe04 WHERE fecha_fin > '2025-12-31'"))
            debug_data['fichas_pasan_2026'] = result.fetchone()[0]
            
            # Fichas TECNÓLOGO
            result = conn.execute(text("SELECT COUNT(*) as cnt FROM programas_formacion_seguimiento_pe04 WHERE nivel_formacion = 'TECNÓLOGO'"))
            debug_data['fichas_tecnologos'] = result.fetchone()[0]
            
            # Fichas PRESENCIAL
            result = conn.execute(text("SELECT COUNT(*) as cnt FROM programas_formacion_seguimiento_pe04 WHERE tipo_formacion = 'PRESENCIAL'"))
            debug_data['fichas_presencial'] = result.fetchone()[0]
            
            # Fichas que cumplen TODAS las condiciones
            result = conn.execute(text("""
                SELECT COUNT(*) as cnt 
                FROM programas_formacion_seguimiento_pe04 
                WHERE nivel_formacion = 'TECNÓLOGO'
                  AND tipo_formacion = 'PRESENCIAL'
                  AND fecha_fin > '2025-12-31'
            """))
            debug_data['fichas_todas_condiciones'] = result.fetchone()[0]
            
            # Ejemplos de registros que cumplen todas las condiciones
            result = conn.execute(text("""
                SELECT numero_ficha, centro_formacion, nivel_formacion, tipo_formacion, fecha_fin, clasificacion_programa_especial, aprendices_activos
                FROM programas_formacion_seguimiento_pe04 
                WHERE nivel_formacion = 'TECNÓLOGO'
                  AND tipo_formacion = 'PRESENCIAL'
                  AND fecha_fin > '2025-12-31'
                LIMIT 5
            """))
            debug_data['ejemplo_registros_todas_condiciones'] = [
                {
                    'numero_ficha': row[0],
                    'centro': row[1],
                    'nivel': row[2],
                    'tipo': row[3],
                    'fecha_fin': row[4].isoformat() if hasattr(row[4], 'isoformat') else str(row[4]),
                    'clasificacion': row[5],
                    'aprendices': row[6]
                }
                for row in result.fetchall()
            ]
            
            return JSONResponse(content=debug_data)
    
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return JSONResponse({'error': str(e), 'detail': error_detail}, status_code=500)
